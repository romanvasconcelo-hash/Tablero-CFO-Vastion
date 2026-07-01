# ============================================================================
# TABLERO CFO VASTION · App de Roberto  v0.3
# Carga + validación + REPORTE INTERNO de indicadores.
# Tras validar, muestra 3 indicadores clave (uno por eje), el P&L de Gestión
# y los indicadores de apoyo. Vista interna (Roberto/Vastion): muestra todo.
# ============================================================================
import streamlit as st
import psycopg2
from psycopg2.extras import execute_values
import xml.etree.ElementTree as ET
import openpyxl, hashlib, io
from datetime import datetime, date, timedelta

st.set_page_config(page_title="Tablero CFO Vastion", page_icon="📊", layout="centered")

def get_conn():
    s = st.secrets["db"]
    return psycopg2.connect(host=s["host"], port=s["port"], dbname=s["dbname"],
                            user=s["user"], password=s["password"])

# ---------------------------------------------------------------------------
# PARSERS
# ---------------------------------------------------------------------------
def _ns(root): return root.tag[root.tag.find('{')+1:root.tag.find('}')]
def _sha(data): return hashlib.sha256(data).hexdigest()

def clasificar(nombre, data):
    if nombre.lower().endswith('.xml'):
        root = ET.fromstring(data); tag = root.tag.lower()
        if 'balanza' in tag: return 'BALANZA'
        if 'catalogo' in tag or 'catálogo' in tag: return 'CATALOGO'
        return 'XML_DESCONOCIDO'
    if nombre.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True); s = wb.sheetnames[0].upper(); wb.close()
        if 'AUXILIAR' in s: return 'AUX_CLIENTES'
        if 'INGRESO' in s: return 'CFDI_EMITIDO'
        if 'EGRESO' in s:  return 'CFDI_RECIBIDO'
        return 'XLSX_DESCONOCIDO'
    return 'DESCONOCIDO'

def parse_catalogo(data):
    root = ET.fromstring(data); ns = {'x': _ns(root)}
    return [dict(num_cuenta=c.attrib['NumCta'], descripcion=c.attrib.get('Desc'),
        cod_agrupador=c.attrib.get('CodAgrup'), subcuenta_de=c.attrib.get('SubCtaDe'),
        nivel=int(c.attrib['Nivel']) if c.attrib.get('Nivel') else None,
        naturaleza=c.attrib.get('Natur')) for c in root.findall('x:Ctas', ns)]

def parse_balanza(data):
    root = ET.fromstring(data); ns = {'x': _ns(root)}
    meta = dict(rfc=root.attrib.get('RFC'), anio=root.attrib.get('Anio'),
                mes=root.attrib.get('Mes'), envio=root.attrib.get('TipoEnvio'))
    rows = [(c.attrib['NumCta'], float(c.attrib['SaldoIni']), float(c.attrib['Debe']),
             float(c.attrib['Haber']), float(c.attrib['SaldoFin'])) for c in root.findall('x:Ctas', ns)]
    return meta, rows


def parse_aux_clientes(data):
    """Auxiliar de la cuenta de Clientes (105). Cada movimiento se autodescribe (col5=cuenta, col4=tercero).
       Devuelve (ejercicio, saldos_ini, movs). saldos_ini: (subcuenta, tercero, saldo_ini).
       movs: (subcuenta, tercero, folio, fecha, tipo_poliza, documento, debe, haber, notas).
       Solo cuentas que abren con '105'. Tipos reales: 'Factura' (cargo) y 'Movimiento Conciliado' (cobro)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]; rows = list(ws.iter_rows(values_only=True)); wb.close()
    ejercicio = None
    for r in rows[:6]:
        for c in r:
            if c and 'Período' in str(c):
                toks = ''.join(ch if ch.isdigit() else ' ' for ch in str(c)).split()
                yrs = [int(t) for t in toks if len(t) == 4]
                if yrs: ejercicio = yrs[0]
    saldos = []; movs = []; i = 0
    while i < len(rows):
        r = rows[i]; c0 = str(r[0]).strip() if r[0] is not None else ''
        if c0 == 'Cuenta:' and i + 1 < len(rows):
            acc = str(rows[i+1][0]) if rows[i+1][0] else ''
            if acc.startswith('105'):
                sub = acc.split('~')[0].strip()
                ter = acc.split('~')[1].strip()[:80] if '~' in acc else None
                saldos.append((sub, ter, _mnum(rows[i+1][1])))
            i += 2; continue
        tp = str(r[2]).strip() if (len(r) > 2 and r[2] is not None) else ''
        if tp in ('Factura', 'Movimiento Conciliado'):
            cta = str(r[5]) if (len(r) > 5 and r[5]) else ''
            if cta.startswith('105'):
                sub = cta.split('~')[0].strip()
                ter = str(r[4])[:80] if (len(r) > 4 and r[4]) else None
                movs.append((sub, ter, str(r[0])[:40] if r[0] else None, _fecha(r[1]), tp,
                             str(r[3])[:40] if (len(r) > 3 and r[3]) else None,
                             _mnum(r[6]), _mnum(r[7]),
                             str(r[10])[:200] if (len(r) > 10 and r[10]) else None))
        i += 1
    if ejercicio is None and movs:
        yrs = [m[3].year for m in movs if m[3]]
        ejercicio = max(yrs) if yrs else None
    return ejercicio, saldos, movs


def _ingest_aux_clientes(cur, cli, data):
    """Parsea y reemplaza el auxiliar de Clientes de su ejercicio (replace-on-load). Devuelve (ejercicio, n_movs)."""
    ejx, saldos_ax, movs_ax = parse_aux_clientes(data)
    if not ejx: return None, 0
    cur.execute("DELETE FROM raw_aux_clientes WHERE cliente_id=%s AND ejercicio=%s", (cli, ejx))
    filas_ax = []
    for sub, ter, sini in saldos_ax:
        sini = sini or 0.0
        filas_ax.append((cli, ejx, sub, ter, None, date(ejx,1,1), 'SALDO INICIAL', None,
                         sini if sini > 0 else 0.0, -sini if sini < 0 else 0.0, None, True))
    for (sub, ter, folio, fecha, tp, doc, debe, haber, notas) in movs_ax:
        filas_ax.append((cli, ejx, sub, ter, folio, fecha, tp, doc, debe or 0.0, haber or 0.0, notas, False))
    if filas_ax:
        execute_values(cur, """INSERT INTO raw_aux_clientes
            (cliente_id,ejercicio,subcuenta,tercero,folio,fecha,tipo_poliza,documento,debe,haber,notas,es_saldo_ini)
            VALUES %s""", filas_ax)
    return ejx, len(movs_ax)


def cargar_auxiliares(cli, lista):
    """Carga uno o varios auxiliares de Clientes de forma independiente (sin balanza).
       lista: [(nombre, data)]. Devuelve [(nombre, ejercicio, n_movs)]."""
    conn = get_conn(); conn.autocommit = False; cur = conn.cursor()
    res = []
    try:
        for nom, data in lista:
            ej, n = _ingest_aux_clientes(cur, cli, data)
            res.append((nom, ej, n))
        conn.commit()
        return res
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

def cargar_cfdi(cli, lista):
    """Carga uno o varios CFDI (emitidos/recibidos) de forma independiente, SIN balanza.
       Inserta filas nuevas completas (Clave SAT y campos de pago incluidos). No modifica lo ya cargado:
       raw_cfdi es inmutable (R-DAT-01). parse_cfdi detecta la direccion por el nombre de la hoja.
       lista: [(nombre, data)]. Devuelve [(nombre, periodo, n_filas)]."""
    conn = get_conn(); conn.autocommit = False; cur = conn.cursor()
    res = []
    try:
        for nom, data in lista:
            pr, n = _ingest_cfdi(cur, cli, nom, data)
            res.append((nom, pr, n))
        conn.commit()
        return res
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

def _num(v):
    try: return float(v) if v not in (None, '') else None
    except: return None
def _mnum(v):
    """Monto que puede venir como '$1,234.56' (auxiliar). Devuelve 0.0 si no parsea."""
    if isinstance(v, (int, float)): return float(v)
    if v in (None, ''): return 0.0
    try: return float(str(v).replace('$', '').replace(',', '').strip())
    except: return 0.0
def _fecha(v):
    if isinstance(v, (datetime, date)): return v
    if not v: return None
    s = str(v)[:19]
    for f in ('%d-%m-%Y','%Y-%m-%d %H:%M:%S','%Y-%m-%dT%H:%M:%S','%Y-%m-%d','%d/%m/%Y'):
        try: return datetime.strptime(s, f)
        except: pass
    return None
def _col(h, n):
    for i,c in enumerate(h):
        if c and n.lower()==str(c).strip().lower(): return i
    for i,c in enumerate(h):
        if c and n.lower() in str(c).strip().lower(): return i

def parse_cfdi(data):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]; direc = 'EMITIDO' if 'INGRESO' in wb.sheetnames[0].upper() else 'RECIBIDO'
    rows = list(ws.iter_rows(values_only=True)); h = rows[3]; wb.close()
    C = {k:_col(h,v) for k,v in dict(rfc='RFC',nom='Razón social',ftim='Fecha Timbrado',
        fexp='Fecha Expedición',uuid='UUID',est='Estatus Sat',tipo='Tipo',uso='Uso CFDI',
        prod='Producto',sub='Subtotal sin descuentos',desc='Descuento',iva16='IVA 16%',iva8='IVA 8%',
        ivar='IVA Retención',isrr='ISR Retención',total='Total',cc='Cuenta Contable',
        cco='Centro de Costos',uuidr='UUID Relacionado',
        metodo='Método de Pago',forma='Forma de Pago',pagado='Pagado',
        pend='Importe pendiente',fpago='Fecha de pago',clave='Clave SAT').items()}
    out = []; cont = {}
    g = lambda r, k: (r[C[k]] if C.get(k) is not None else None)   # columna ausente -> None (CFDI recibidos)
    for r in rows[4:]:
        u = r[C['uuid']]
        if not u: continue
        cont[u] = cont.get(u,0)+1
        cc = r[C['cc']]; acc = str(cc).split(' - ')[0].strip() if cc else None
        fx = _fecha(r[C['fexp']]); per = date(fx.year, fx.month, 1) if fx else None
        pgo = g(r,'pagado'); pagado = (str(pgo).strip().upper()=='SI') if pgo is not None else None
        out.append((u,cont[u],direc,per,fx,_fecha(r[C['ftim']]),r[C['tipo']],r[C['uso']],r[C['est']],
            r[C['rfc']], str(r[C['nom']])[:120] if r[C['nom']] else None,
            str(r[C['prod']])[:200] if r[C['prod']] else None,
            _num(r[C['sub']]),_num(r[C['desc']]),_num(r[C['iva16']]),_num(r[C['iva8']]),
            _num(r[C['ivar']]),_num(r[C['isrr']]),_num(r[C['total']]),acc,r[C['cco']],r[C['uuidr']],
            g(r,'metodo'), g(r,'forma'), pagado, _num(g(r,'pend')), _fecha(g(r,'fpago')),
            (str(g(r,'clave')).strip()[:20] if g(r,'clave') is not None else None)))
    return out

# ---------------------------------------------------------------------------
# CARGA + VALIDACIÓN
# ---------------------------------------------------------------------------
def registrar(cur, cli, per, tipo, envio, nombre, data):
    cur.execute("""INSERT INTO origen_archivo (cliente_id,periodo,tipo,envio,storage_path,hash_sha256,bytes,cargado_por)
        VALUES (%s,%s,%s,%s,%s,%s,%s,'streamlit') ON CONFLICT (cliente_id,periodo,tipo,hash_sha256) DO NOTHING RETURNING id""",
        (cli, per, tipo, envio, nombre, _sha(data), len(data)))
    row = cur.fetchone()
    if row: return row[0], True
    cur.execute("SELECT id FROM origen_archivo WHERE cliente_id=%s AND periodo=%s AND tipo=%s AND hash_sha256=%s",
                (cli, per, tipo, _sha(data))); return cur.fetchone()[0], False

def _ingest_cfdi(cur, cli, nombre, data):
    """Parsea e ingesta un CFDI (emitido/recibido) en su PRIMERA carga, completo: Clave SAT y campos de pago
       entran en el mismo INSERT. ON CONFLICT DO NOTHING respeta la inmutabilidad de raw_cfdi (R-DAT-01,
       trigger tg_cfdi_inmutable): no se modifica lo ya cargado. Devuelve (periodo, n_filas)."""
    crows = parse_cfdi(data)
    if not crows:
        return None, 0
    pr = crows[0][3]
    arch, _new = registrar(cur, cli, pr, 'CFDI', None, nombre, data)
    execute_values(cur, """INSERT INTO raw_cfdi (uuid,renglon,direccion,periodo,fecha_emision,fecha_timbrado,tipo_cfdi,uso_cfdi,estatus_sat,contraparte_rfc,contraparte_nom,concepto,subtotal,descuento,iva_16,iva_8,iva_retenido,isr_retenido,total,cuenta_contable,centro_costos,uuid_relacionado,metodo_pago,forma_pago,pagado,importe_pendiente,fecha_pago,clave_sat,cliente_id,archivo_id)
        VALUES %s ON CONFLICT DO NOTHING""",
        [t+(cli,arch) for t in crows])
    return pr, len(crows)


def procesar(cli, archivos):
    conn = get_conn(); conn.autocommit = False; cur = conn.cursor()
    try:
        if 'CATALOGO' in archivos:
            nom, data = archivos['CATALOGO']; cat_rows = parse_catalogo(data)
        nomb, datab = archivos['BALANZA']
        meta, brows = parse_balanza(datab); per = f"{meta['anio']}-{meta['mes']}-01"
        envio = 'COMPLEMENTARIA' if meta['envio']=='C' else 'NORMAL'
        cur.execute("UPDATE cliente SET rfc=%s WHERE id=%s AND (rfc IS NULL OR rfc='')", (meta['rfc'], cli))
        if 'CATALOGO' in archivos:
            cat_id, cat_new = registrar(cur, cli, per, 'CATALOGO', None, nom, data)
            if cat_new:
                execute_values(cur, """INSERT INTO raw_catalogo (archivo_id,cliente_id,num_cuenta,descripcion,cod_agrupador,subcuenta_de,nivel,naturaleza)
                    VALUES %s ON CONFLICT DO NOTHING""",
                    [(cat_id,cli,c['num_cuenta'],c['descripcion'],c['cod_agrupador'],c['subcuenta_de'],c['nivel'],c['naturaleza']) for c in cat_rows])
        cur.execute("""SELECT id FROM origen_archivo WHERE cliente_id=%s AND tipo='CATALOGO' AND periodo<=%s
                       ORDER BY periodo DESC,version DESC LIMIT 1""", (cli, per))
        rcat = cur.fetchone()
        if not rcat: raise RuntimeError("No hay catálogo para este cliente. Sube el catálogo XML la primera vez.")
        cat_id = rcat[0]
        bal_id, bal_new = registrar(cur, cli, per, 'BALANZA', envio, nomb, datab)
        if bal_new:
            execute_values(cur, "INSERT INTO raw_balanza (archivo_id,num_cuenta,saldo_inicial,debe,haber,saldo_final) VALUES %s ON CONFLICT DO NOTHING",
                [(bal_id,)+r for r in brows])
        cur.execute("DELETE FROM insumos_balanza WHERE archivo_id=%s", (bal_id,))
        cur.execute("""INSERT INTO insumos_balanza (archivo_id,cliente_id,periodo,num_cuenta,cod_agrupador,naturaleza,es_hoja,es_orden,bloque,es_laboral,saldo_final)
            SELECT b.archivo_id,%(cli)s,%(per)s,b.num_cuenta,c.cod_agrupador,c.naturaleza,
              (b.num_cuenta NOT IN (SELECT subcuenta_de FROM raw_catalogo WHERE archivo_id=%(cat)s AND subcuenta_de IS NOT NULL)),
              (b.num_cuenta LIKE '8%%'),fb.bloque,fb.es_laboral,b.saldo_final
            FROM raw_balanza b JOIN raw_catalogo c ON c.archivo_id=%(cat)s AND c.num_cuenta=b.num_cuenta
            CROSS JOIN LATERAL fn_bloque(%(cli)s,c.cod_agrupador) fb WHERE b.archivo_id=%(bal)s""",
            dict(cli=cli, per=per, cat=cat_id, bal=bal_id))
        cur.execute("""INSERT INTO periodo_estado (cliente_id,periodo,estado,archivo_vigente) VALUES (%s,%s,'RECIBIDO',%s)
                       ON CONFLICT (cliente_id,periodo) DO UPDATE SET archivo_vigente=EXCLUDED.archivo_vigente""", (cli, per, bal_id))
        for tipo in ('CFDI_EMITIDO', 'CFDI_RECIBIDO'):
            if tipo not in archivos: continue
            nomc, datac = archivos[tipo]
            _ingest_cfdi(cur, cli, nomc, datac)
        if 'AUX_CLIENTES' in archivos:
            _ingest_aux_clientes(cur, cli, archivos['AUX_CLIENTES'][1])
        cur.execute("SELECT prueba,paso,severidad,detalle FROM fn_validar_periodo(%s)", (bal_id,)); integ = cur.fetchall()
        cur.execute("SELECT prueba,paso,severidad,detalle FROM fn_validar_madurez(%s)", (bal_id,)); madz = cur.fetchall()
        bloqueo = any((not ok) and sev=='BLOQUEANTE' for _,ok,sev,_ in integ) or any((not ok) and sev=='BLOQUEANTE' for _,ok,sev,_ in madz)
        if not bloqueo:
            cur.execute("UPDATE periodo_estado SET estado='VALIDADO',validado_en=now() WHERE cliente_id=%s AND periodo=%s", (cli, per))
        conn.commit()
        return per, bal_id, integ, madz
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

# ---------------------------------------------------------------------------
# INDICADORES (reporte interno)
# ---------------------------------------------------------------------------
def cargar_indicadores(bal_id, cli, per):
    conn = get_conn(); cur = conn.cursor(); ind = {}
    try:
        cur.execute("SELECT concepto,monto,pct FROM fn_pl_gestion(%s) ORDER BY orden", (bal_id,))
        pl = cur.fetchall()
        ind['pl'] = [(c, float(m), (float(p) if p is not None else None)) for c,m,p in pl]
        d = {c:(float(m), (float(p) if p is not None else None)) for c,m,p in pl}
        ing = d.get('Ingresos',(0,0))[0]
        lab = d.get('(-) Eficiencia Laboral',(0,0))[0]
        mb  = d.get('= MARGEN BRUTO DE GESTIÓN',(0,0))[0]
        ind['pretax_pct'] = d.get('= UTILIDAD ANTES DE IMPUESTOS',(0,None))[1]
        ind['mb_pct']     = d.get('= MARGEN BRUTO DE GESTIÓN',(0,None))[1]
        ind['nomina_pct'] = round(lab/ing*100,1) if ing else None
        ind['gpld']       = round(mb/lab,2) if lab and mb>0 else None
        cur.execute("SELECT concepto,valor FROM fn_cash_lag(%s)", (bal_id,))
        cl = {c:(float(v) if v is not None else None) for c,v in cur.fetchall()}
        ind['cash_lag'] = next((v for k,v in cl.items() if k.startswith('Cash Lag')), None)
        ind['caja']     = cl.get('Caja fin de mes')
        cur.execute("SELECT concepto,valor FROM fn_eiva(%s,%s)", (cli, per))
        ev = {c:(float(v) if v is not None else None) for c,v in cur.fetchall()}
        ind['eiva_pct']  = ev.get('EIVA % (acred/tras)')
        ind['iva_neto']  = ev.get('IVA neto (+cargo / -favor)')
        return ind
    finally:
        cur.close(); conn.close()

def sem_pretax(p):
    if p is None: return "—"
    if p >= 10: return "🟢 Sano (≥10%)"
    if p >= 5:  return "🟡 Mínimo (5–9%)"
    return "🔴 Peligro (<5%)"

def money(v):
    return "—" if v is None else f"${v:,.0f}"


def _efe_diag(efe):
    """R-EFE-01 -> (cuadra, causa, accion). Traduce el descuadre a la causa raiz y la accion."""
    if efe.get("cuadra", True):
        return True, "", ""
    plug = efe.get("plug", 0.0); apd = efe.get("ap_descuadre", 0.0); apr = efe.get("ap_resultado", 0.0)
    if not efe.get("base_anual", False):
        return False, "Flujo mensual sin base anual: falta enero del ejercicio para evaluar la apertura.", \
               "Cargar la balanza de enero del ejercicio."
    if abs(apr) >= 1:
        return False, "Ejercicio anterior sin cerrar: las cuentas de resultados abren en " + _fmt(apr) + " (deben abrir en cero).", \
               "Cerrar el ejercicio anterior (llevar el resultado a ejercicios anteriores) antes de reportar."
    if abs(apd) >= 1:
        return False, "La apertura del ejercicio no cuadra por " + _fmt(apd) + " (el balance inicial no suma cero).", \
               "Revisar la captura/cierre de la apertura del ejercicio."
    return False, "El flujo no cuadra por " + _fmt(plug) + " con apertura correcta: un agrupador mueve efectivo sin origen.", \
           "Revisar la clasificacion de agrupadores de financiamiento/inversion."


# ---------------------------------------------------------------------------
# REPORTE INTERNO · Comportamiento del ejercicio (tendencia + resumen)
# ---------------------------------------------------------------------------
# Spec unica: alimenta la grafica y la tabla resumen (DRY). meta y dir solo en anclas duras.
SERIE_SPEC = [
    ("Rentabilidad (antes de impuestos) %", "pretax", "pct",     10.0, "mayor"),   # Crabtree
    ("Margen bruto %",           "mb",     "pct",     None, None),       # segun sector
    ("Liquidez (AC / PC)",       "ac_pc",  "x",       1.1,  "mayor"),    # licitacion
    ("Capital de trabajo",       "cnt",    "money",   0.0,  "mayor"),    # piso: positivo
    ("Caja al cierre",           "caja",   "money",   None, None),
    ("Endeudamiento (PT / AT)",  "pt_at",  "pctfrac", 0.70, "menor"),    # licitacion
]

def serie_anual(cli, anio):
    """Recorre los meses cargados del ejercicio y arma la serie de indicadores clave."""
    import datetime as _dt
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT periodo, archivo_vigente FROM periodo_estado
                       WHERE cliente_id=%s AND archivo_vigente IS NOT NULL
                         AND periodo >= %s AND periodo <= %s ORDER BY periodo""",
                    (cli, _dt.date(anio,1,1), _dt.date(anio,12,1)))
        rows = cur.fetchall()
    finally:
        cur.close(); conn.close()
    serie = []
    for per, arch in rows:
        ef = estados_financieros(arch); s = _stocks_de(ef); y = ef["ytd"]
        ing = y["ing"]
        serie.append(dict(
            mes=str(per)[:7],
            pretax=(y["uai"]/ing*100) if ing else None,
            mb=(y["ub"]/ing*100) if ing else None,
            ac_pc=(s["act_circ"]/s["pas_circ"]) if s["pas_circ"] else None,
            cnt=s["act_circ"]-s["pas_circ"],
            caja=ef["efe"]["efec_fin"],
            pt_at=(s["pasivo"]/s["activo"]) if s["activo"] else None,
        ))
    return serie

def _fig_tendencia(serie):
    """Panel 2x3 de tendencia del ejercicio -> PNG bytes. Linea de meta punteada donde hay ancla."""
    import matplotlib; matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import io as _io
    DARK="#1c2d3a"; ACC="#2c3e50"; RED="#c0392b"
    meses = [d["mes"][5:7] for d in serie]; x = list(range(len(serie)))
    fig, axes = plt.subplots(2, 3, figsize=(11, 6))
    for ax, (titulo, key, fmt, meta, _dir) in zip(axes.flat, SERIE_SPEC):
        ys = [d[key] for d in serie]
        yv = [(float("nan") if v is None else (v/1e6 if fmt=="money" else (v*100 if fmt=="pctfrac" else v))) for v in ys]
        ax.plot(x, yv, marker="o", ms=3, lw=1.6, color=DARK)
        if meta is not None:
            ml = meta/1e6 if fmt=="money" else (meta*100 if fmt=="pctfrac" else meta)
            ax.axhline(ml, ls="--", lw=1.0, color=RED)
        suf = " ($M)" if fmt=="money" else ("" )
        ax.set_title(titulo+suf, fontsize=9, color=ACC, loc="left")
        ax.set_xticks(x); ax.set_xticklabels(meses, fontsize=6)
        ax.tick_params(axis="y", labelsize=6)
        for sp in ("top","right"): ax.spines[sp].set_visible(False)
        ax.grid(axis="y", lw=0.3, alpha=0.4)
    fig.tight_layout(pad=1.2)
    buf = _io.BytesIO(); fig.savefig(buf, format="png", dpi=150); plt.close(fig)
    return buf.getvalue()

def _fserie(v, fmt):
    if v is None: return "-"
    if fmt=="pct":     return "{:.1f}%".format(v)
    if fmt=="x":       return "{:.2f}x".format(v)
    if fmt=="money":   return _fmt(v)
    if fmt=="pctfrac": return "{:.1f}%".format(v*100)
    return "{:.2f}".format(v)

def _dias_ytd(per):
    import datetime
    y, m = int(per[:4]), int(per[5:7])
    d = datetime.date(y,12,31) if m==12 else datetime.date(y,m+1,1)-datetime.timedelta(days=1)
    return d.timetuple().tm_yday

def cierre_ejercicio(cli, anio):
    """Ultimo mes cargado del ejercicio -> (periodo 'YYYY-MM', archivo_vigente). None si no hay."""
    import datetime as _dt
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT periodo, archivo_vigente FROM periodo_estado
                       WHERE cliente_id=%s AND archivo_vigente IS NOT NULL
                         AND periodo>=%s AND periodo<=%s ORDER BY periodo DESC LIMIT 1""",
                    (cli, _dt.date(anio,1,1), _dt.date(anio,12,1)))
        r = cur.fetchone()
        return (str(r[0])[:7], r[1]) if r else None
    finally:
        cur.close(); conn.close()

def _dias_mes(periodo):
    import datetime as _dt
    y, m = int(periodo[:4]), int(periodo[5:7])
    return (_dt.date(y,12,31) if m==12 else _dt.date(y,m+1,1)-_dt.timedelta(days=1)).day

def base_poder_uno(ef, ef_prev, periodo, temporalidad):
    """Base de ingresos/costo/gastos y stocks para el Poder del Uno, segun la temporalidad elegida."""
    ytd = ef["ytd"]; s = _stocks_de(ef); dy = _dias_ytd(periodo)
    if temporalidad == "mensual":
        p = ef_prev["ytd"] if ef_prev else {}
        ing = ytd["ing"]-p.get("ing",0); cos = ytd["cos"]-p.get("cos",0); gas = ytd["gas"]-p.get("gas",0)
        dias = _dias_mes(periodo)
    elif temporalidad == "anualizado":
        f = 365/dy if dy else 1
        ing = ytd["ing"]*f; cos = ytd["cos"]*f; gas = ytd["gas"]*f; dias = 365
    else:  # ytd
        ing = ytd["ing"]; cos = ytd["cos"]; gas = ytd["gas"]; dias = dy
    return dict(ing=ing, cos=cos, gas=gas, cxc=s["cxc"], inv=s["inv"], cxp=s["cxp"], dias=dias)

def poder_uno_tabla(base, mv):
    """Impacto en pesos de cada palanca al movimiento capturado. mv: precio/volumen/costo/gastos (%) y cxc/inv/cxp (dias).
    Devuelve (filas, delta_utilidad, delta_caja, trampa_volumen, utilidad_bruta)."""
    ing, cos, gas, dias = base["ing"], base["cos"], base["gas"], base["dias"]
    ub = ing - cos; rev = ing/dias if dias else 0; cd = cos/dias if dias else 0
    filas = [
        ("Precio",              "+{:.1f}%".format(mv["precio"]),  ing*mv["precio"]/100,  "utilidad"),
        ("Volumen",             "+{:.1f}%".format(mv["volumen"]), ub*mv["volumen"]/100,  "utilidad"),
        ("Costo de ventas",     "-{:.1f}%".format(mv["costo"]),   cos*mv["costo"]/100,   "utilidad"),
        ("Gastos de operacion", "-{:.1f}%".format(mv["gastos"]),  gas*mv["gastos"]/100,  "utilidad"),
        ("Dias por cobrar",     "-{:.0f} dias".format(mv["cxc"]), rev*mv["cxc"],         "caja"),
        ("Dias de inventario",  "-{:.0f} dias".format(mv["inv"]), cd*mv["inv"],          "caja"),
        ("Dias por pagar",      "+{:.0f} dias".format(mv["cxp"]), cd*mv["cxp"],          "caja"),
    ]
    du = sum(v for _,_,v,k in filas if k=="utilidad")
    dc = sum(v for _,_,v,k in filas if k=="caja")
    return filas, du, dc, (ub < 0 and mv["volumen"] > 0), ub

def modelo_negocio(ef, dias):
    """Tabla 3.10 de Alexander con los indicadores que el sistema ya calcula (cierre del ejercicio)."""
    y = ef["ytd"]; s = _stocks_de(ef); ing = y["ing"]
    sd = lambda a,b: (a/b) if b else None
    un = y.get("un", y["uai"]); isr = y.get("isr_prov", 0.0)
    prof = [
        ("Ventas", ing, 1.0),
        ("(-) Costo de ventas", y["cos"], sd(y["cos"], ing)),
        ("= Margen bruto", y["ub"], sd(y["ub"], ing)),
        ("(-) Gastos de operacion", y["gas"], sd(y["gas"], ing)),
        ("(-) Depreciacion", y["dep"], sd(y["dep"], ing)),
        ("= Utilidad de operacion", y["ebit"], sd(y["ebit"], ing)),
        ("(-) Resultado financiero neto", y["fin"], sd(y["fin"], ing)),
        ("(-) ISR provisional", isr, sd(isr, ing)),
        ("= Utilidad neta", un, sd(un, ing)),
    ]
    dso = sd(s["cxc"], ing); dio = sd(s["inv"], y["cos"])
    asset = [
        ("Dias cuentas por cobrar", (dso*dias if dso is not None else None), "dias"),
        ("Dias de inventario", (dio*dias if dio is not None else None), "dias"),
        ("Rotacion de capital operativo", sd(ing, ef["cash"]["opcap"]), "x"),
        ("Rotacion de activo fijo", sd(ing, s["afn"]), "x"),
        ("Rotacion de activo total", sd(ing, s["activo"]), "x"),
        ("Cobertura de intereses", sd(y["ebit"], y["fin"]), "x"),
    ]
    lev = [("Deuda / Capital total", sd(s["pasivo"], s["pasivo"]+s["capital"]), "pct")]
    ret = [("Rendimiento sobre capital", sd(un, s["capital"]), "pct"),
           ("Rendimiento sobre capital empleado", sd(y["ebit"], s["activo"]-s["pas_circ"]), "pct")]
    return dict(prof=prof, asset=asset, lev=lev, ret=ret)

def _bmval(v, fmt):
    if v is None: return "-"
    if fmt == "dias": return "{:,.0f} dias".format(v)
    if fmt == "x":    return "{:.2f}x".format(v)
    if fmt == "pct":  return "{:.1f}%".format(v*100)
    return "{:.2f}".format(v)

CSF_CAMPOS = ["rfc","razon_social","nombre_comercial","regimen_capital","actividad_economica",
              "actividad_pct","regimen_fiscal","fecha_inicio_ops","estatus","cp","municipio","entidad"]

def parse_constancia(pdf_bytes):
    """Extrae datos clave de una Constancia de Situacion Fiscal (SAT). Devuelve dict con CSF_CAMPOS."""
    import pdfplumber, io, re
    out = {k: "" for k in CSF_CAMPOS}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        p1 = pdf.pages[0].extract_text(x_tolerance=1) or ""
        p2 = pdf.pages[1].extract_text(x_tolerance=1) if len(pdf.pages) > 1 else ""
    def g(pat, s, grp=1, flags=0):
        m = re.search(pat, s, flags); return m.group(grp).strip() if m else ""
    out["rfc"]              = g(r'RFC:\s*([A-Z0-9&\u00d1]{12,13})', p1)
    out["razon_social"]     = g(r'Denominaci[o\u00f3]n/Raz[o\u00f3]n Social:\s*(.+)', p1)
    out["regimen_capital"]  = g(r'R[e\u00e9]gimen Capital:\s*(.+)', p1)
    out["nombre_comercial"] = g(r'Nombre Comercial:\s*(.+)', p1)
    out["fecha_inicio_ops"] = g(r'Fecha inicio de operaciones:\s*(.+)', p1)
    out["estatus"]          = g(r'Estatus en el padr[o\u00f3]n:\s*(.+)', p1)
    out["cp"]               = g(r'C[o\u00f3]digo Postal:\s*(\d{5})', p1)
    out["municipio"]        = g(r'Municipio o Demarcaci[o\u00f3]n Territorial:\s*(.+)', p1)
    out["entidad"]          = g(r'Entidad Federativa:\s*(.+?)(?:\s+Entre Calle:|$)', p1)
    out["actividad_economica"] = g(r'^\s*\d+\s+(.+?)\s+\d+\s+\d{2}/\d{2}/\d{4}', p2, flags=re.M)
    out["actividad_pct"]       = g(r'^\s*\d+\s+.+?\s+(\d+)\s+\d{2}/\d{2}/\d{4}', p2, flags=re.M)
    mreg = re.search(r'R[e\u00e9]gimen Fecha Inicio.*?\n(.+?)\s+\d{2}/\d{2}/\d{4}', p2, re.S)
    out["regimen_fiscal"] = mreg.group(1).strip() if mreg else ""
    return out

def get_cliente_csf(cli):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT " + ",".join(CSF_CAMPOS) + " FROM cliente_csf WHERE cliente_id=%s", (cli,))
        r = cur.fetchone()
        return dict(zip(CSF_CAMPOS, r)) if r else None
    except Exception:
        return None
    finally:
        cur.close(); conn.close()

def save_cliente_csf(cli, data):
    conn = get_conn(); cur = conn.cursor()
    try:
        cols = ",".join(CSF_CAMPOS); ph = ",".join(["%s"] * len(CSF_CAMPOS))
        upd = ",".join(c + "=EXCLUDED." + c for c in CSF_CAMPOS)
        cur.execute("INSERT INTO cliente_csf (cliente_id," + cols + ",actualizado) "
                    "VALUES (%s," + ph + ",now()) "
                    "ON CONFLICT (cliente_id) DO UPDATE SET " + upd + ", actualizado=now()",
                    [cli] + [(data.get(c) or "") for c in CSF_CAMPOS])
        conn.commit()
    finally:
        cur.close(); conn.close()

def _hex_rgb(h, default=(28, 45, 58)):
    try:
        h = (h or "").lstrip("#")
        return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
    except Exception:
        return default

def get_branding(cli):
    """Identidad visual del cliente: color_primario, color_acento, logo (bytes), logo_nombre."""
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT color_primario, color_acento, logo, logo_nombre FROM cliente_branding WHERE cliente_id=%s", (cli,))
        r = cur.fetchone()
        if not r:
            return None
        return dict(color_primario=r[0], color_acento=r[1],
                    logo=(bytes(r[2]) if r[2] is not None else None), logo_nombre=r[3])
    except Exception:
        return None
    finally:
        cur.close(); conn.close()

def save_branding(cli, color_primario, color_acento, logo_bytes=None, logo_nombre=None):
    """Upsert de identidad visual. Si logo_bytes es None, conserva el logo existente (solo actualiza colores)."""
    conn = get_conn(); cur = conn.cursor()
    try:
        if logo_bytes is not None:
            cur.execute("""INSERT INTO cliente_branding (cliente_id,color_primario,color_acento,logo,logo_nombre,actualizado)
                           VALUES (%s,%s,%s,%s,%s,now())
                           ON CONFLICT (cliente_id) DO UPDATE SET color_primario=EXCLUDED.color_primario,
                             color_acento=EXCLUDED.color_acento, logo=EXCLUDED.logo,
                             logo_nombre=EXCLUDED.logo_nombre, actualizado=now()""",
                        (cli, color_primario, color_acento, psycopg2.Binary(logo_bytes), logo_nombre))
        else:
            cur.execute("""INSERT INTO cliente_branding (cliente_id,color_primario,color_acento,actualizado)
                           VALUES (%s,%s,%s,now())
                           ON CONFLICT (cliente_id) DO UPDATE SET color_primario=EXCLUDED.color_primario,
                             color_acento=EXCLUDED.color_acento, actualizado=now()""",
                        (cli, color_primario, color_acento))
        conn.commit()
    finally:
        cur.close(); conn.close()

def _pdf_tabla_estado(pdf, t, Mg, CW, titulo, rows, ca, cp):
    dual = cp is not None
    if pdf.get_y() + 18 > 282: pdf.add_page()
    pdf.set_x(Mg); pdf.set_fill_color(28,45,58); pdf.set_text_color(255,255,255); pdf.set_font("Helvetica","B",9)
    pdf.cell(CW,6,t(" " + titulo), fill=True); pdf.ln(7)
    pdf.set_x(Mg); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","B",7.5)
    if dual:
        pdf.cell(86,5,t("Concepto")); pdf.cell(34,5,t(ca),align="R"); pdf.cell(34,5,t(cp),align="R"); pdf.cell(CW-154,5,t("Variacion"),align="R")
    else:
        pdf.cell(120,5,t("Concepto")); pdf.cell(CW-120,5,t(ca),align="R")
    pdf.ln(5)
    for label, va, vp, bold in rows:
        if pdf.get_y() + 5.2 > 286: pdf.add_page()
        if va is None and vp is None:
            pdf.set_x(Mg); pdf.set_fill_color(238,240,242); pdf.set_text_color(80,90,100); pdf.set_font("Helvetica","B",7.5)
            pdf.cell(CW,5,t(" " + label), fill=True); pdf.ln(5); continue
        pdf.set_x(Mg); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","B" if bold else "",8)
        if dual:
            pdf.cell(86,5,t(label)); pdf.cell(34,5,t(_fmt(va)),align="R")
            pdf.cell(34,5,t(_fmt(vp) if vp is not None else "-"),align="R")
            pdf.cell(CW-154,5,t(_fmt((va or 0)-(vp or 0))),align="R")
        else:
            pdf.cell(120,5,t(label)); pdf.cell(CW-120,5,t(_fmt(va)),align="R")
        pdf.ln(5)
    pdf.ln(3)

def _pdf_tabla_ratios(pdf, t, Mg, CW, a_rows, p_rows, ca, cp):
    pmap = {(s,l):(v,f) for s,l,v,f in p_rows} if p_rows else {}
    dual = bool(p_rows); sec = None
    for s,l,v,f in a_rows:
        if s != sec:
            if pdf.get_y() + 14 > 282: pdf.add_page()
            pdf.set_x(Mg); pdf.set_fill_color(28,45,58); pdf.set_text_color(255,255,255); pdf.set_font("Helvetica","B",8.5)
            pdf.cell(CW,6,t(" " + s), fill=True); pdf.ln(7)
            pdf.set_x(Mg); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","B",7.5)
            if dual:
                pdf.cell(86,5,t("Ratio")); pdf.cell(34,5,t(ca),align="R"); pdf.cell(34,5,t(cp),align="R"); pdf.cell(CW-154,5,t("Variacion"),align="R")
            else:
                pdf.cell(120,5,t("Ratio")); pdf.cell(CW-120,5,t(ca),align="R")
            pdf.ln(5); sec = s
        if pdf.get_y() + 5 > 286: pdf.add_page()
        vp = pmap.get((s,l),(None,f))[0]
        pdf.set_x(Mg); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","",8)
        if dual:
            pdf.cell(86,5,t(l)); pdf.cell(34,5,t(_fr(v,f)),align="R")
            pdf.cell(34,5,t(_fr(vp,f)),align="R"); pdf.cell(CW-154,5,t(_fvar(v,vp,f)),align="R")
        else:
            pdf.cell(120,5,t(l)); pdf.cell(CW-120,5,t(_fr(v,f)),align="R")
        pdf.ln(5)
    pdf.ln(3)

def pdf_interno_2025(nombre, anio, serie, ef_cierre=None, dias_cierre=365, crecimiento=None,
                     ef_cierre_prev=None, rat_now=None, rat_prev=None, datos_cliente=None, fiscal_sem=None):
    """Reporte interno: modelo de negocio (cierre) + tendencia y resumen del ejercicio."""
    from fpdf import FPDF
    import io as _io
    class _PDFInt(FPDF):
        def footer(self):
            self.set_y(-12); self.set_font("Helvetica","",7); self.set_text_color(150,150,150)
            self.set_x(self.l_margin)
            self.cell(0,5,"Confidencial - uso interno Vastion  |  Pagina " + str(self.page_no()),align="C")
    def t(s):
        return (str(s).replace("\u2014","-").replace("\u2013","-").replace("\u2212","-")
                .replace("\u0394","Var. ")
                .encode("latin-1","replace").decode("latin-1"))
    W, Mg = 210, 12; CW = W - 2*Mg
    pdf = _PDFInt(orientation="P", unit="mm", format="A4"); pdf.set_auto_page_break(True, 15)

    def _band(titulo):
        yb = pdf.get_y()
        pdf.set_fill_color(28,45,58); pdf.rect(Mg,yb,CW,6,style="F")
        pdf.set_xy(Mg+1.5,yb+1); pdf.set_text_color(255,255,255); pdf.set_font("Helvetica","B",8.5)
        pdf.cell(CW-3,4,t(titulo)); pdf.set_y(yb+8)

    # ===== Portada: Datos generales del cliente =====
    if datos_cliente:
        dc = datos_cliente
        pdf.add_page()
        pdf.set_fill_color(28,45,58); pdf.rect(0,0,W,56,style="F")
        pdf.set_text_color(255,255,255); pdf.set_xy(Mg,15); pdf.set_font("Helvetica","B",11)
        pdf.cell(CW,6,t("REPORTE CFO  -  USO INTERNO VASTION"))
        pdf.set_xy(Mg,24); pdf.set_font("Helvetica","B",22)
        pdf.multi_cell(CW,9,t(dc.get("razon_social") or nombre))
        pdf.set_xy(Mg,46); pdf.set_font("Helvetica","",11)
        pdf.cell(CW,6,t("Ejercicio " + str(anio)))
        pdf.set_y(68)
        def _fila(lbl, val):
            if not val: return
            yb = pdf.get_y()
            pdf.set_x(Mg); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","B",9)
            pdf.cell(52,6,t(lbl))
            pdf.set_xy(Mg+52, yb); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","",10)
            pdf.multi_cell(CW-52,6,t(val)); pdf.ln(2)
        _fila("RFC", dc.get("rfc"))
        _fila("Nombre comercial", dc.get("nombre_comercial"))
        _fila("Régimen capital", dc.get("regimen_capital"))
        _act = dc.get("actividad_economica") or ""
        if dc.get("actividad_pct"): _act = (_act + "  (" + str(dc["actividad_pct"]) + "%)").strip()
        _fila("Actividad económica", _act)
        _fila("Régimen fiscal", dc.get("regimen_fiscal"))
        _fila("Estatus en el padrón", dc.get("estatus"))
        _fila("Inicio de operaciones", dc.get("fecha_inicio_ops"))
        _dom = ", ".join([x for x in [dc.get("municipio"), dc.get("entidad"),
                          ("C.P. " + dc["cp"]) if dc.get("cp") else ""] if x])
        _fila("Domicilio fiscal", _dom)
        pdf.ln(4); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","",7.5)
        pdf.multi_cell(CW,3.6,t("Datos tomados de la Constancia de Situacion Fiscal del contribuyente. "
                                "Documento de uso interno; no constituye opinion ni dictamen fiscal."))

    # ===== Pagina 1: Modelo de negocio (cierre del ejercicio) =====
    if ef_cierre is not None:
        mn = modelo_negocio(ef_cierre, dias_cierre)
        pdf.add_page()
        pdf.set_fill_color(28,45,58); pdf.rect(0,0,W,22,style="F")
        pdf.set_text_color(255,255,255); pdf.set_xy(Mg,5); pdf.set_font("Helvetica","B",14)
        pdf.cell(CW,7,t("Modelo de negocio - cierre " + str(anio)))
        pdf.set_xy(Mg,13); pdf.set_font("Helvetica","",9)
        pdf.cell(CW,5,t("Como gana dinero el negocio: rentabilidad, uso de activos, apalancamiento y rendimientos."))
        pdf.set_y(27)
        if crecimiento is not None:
            pdf.set_x(Mg); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","",9)
            pdf.cell(CW,6,t("Crecimiento de ventas vs ejercicio anterior: {:+.1f}%".format(crecimiento*100))); pdf.ln(8)
        _band("RENTABILIDAD  (estado de resultados de gestion)")
        pdf.set_x(Mg); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","B",7.5)
        pdf.cell(96,5,t("Concepto")); pdf.cell(50,5,t("Monto"),align="R"); pdf.cell(CW-146,5,t("% ventas"),align="R"); pdf.ln(5)
        for lbl, monto, pct in mn["prof"]:
            bold = lbl.startswith("=")
            pdf.set_x(Mg); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","B" if bold else "",8)
            pdf.cell(96,5,t(lbl)); pdf.cell(50,5,t(_fmt(monto)),align="R")
            pdf.cell(CW-146,5,t("{:.1f}%".format(pct*100) if pct is not None else "-"),align="R"); pdf.ln(5)
        pdf.ln(2)
        for titulo, rows in [("USO DE ACTIVOS", mn["asset"]),
                             ("APALANCAMIENTO", mn["lev"]),
                             ("RENDIMIENTOS", mn["ret"])]:
            _band(titulo)
            for lbl, val, fmt in rows:
                pdf.set_x(Mg); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","",8)
                pdf.cell(120,5,t(lbl)); pdf.cell(CW-120,5,t(_bmval(val,fmt)),align="R"); pdf.ln(5)
            pdf.ln(2)
        pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","",7.5)
        pdf.multi_cell(CW,3.6,t("Estructura del modelo de negocio poblada con los indicadores del sistema al cierre del "
                                "ejercicio. El modelo explica como esta armado el negocio (margen, rotacion y apalancamiento); "
                                "las metas vienen de fuentes externas, no de esta vista."))

    # ===== Pagina: Comportamiento del ejercicio (tendencia + resumen) =====
    pdf.add_page()
    pdf.set_fill_color(28,45,58); pdf.rect(0,0,W,26,style="F")
    pdf.set_text_color(255,255,255); pdf.set_xy(Mg,7); pdf.set_font("Helvetica","B",16)
    pdf.cell(CW,8,t("Reporte interno - Comportamiento " + str(anio)))
    pdf.set_xy(Mg,16); pdf.set_font("Helvetica","",10)
    pdf.cell(CW,6,t(nombre + "   |   uso interno Vastion   |   acumulado del ejercicio"))
    png = _fig_tendencia(serie)
    pdf.image(_io.BytesIO(png), x=Mg, y=31, w=CW)
    yt = 31 + CW*6.0/11.0 + 6
    pdf.set_xy(Mg, yt); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","B",12)
    pdf.cell(CW,7,t("Resumen del ejercicio")); pdf.ln(9)
    pdf.set_x(Mg); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","B",8)
    for w,txt,al in [(60,"Indicador","L"),(26,"Apertura","R"),(26,"Cierre","R"),(26,"Promedio","R"),(22,"Meta","R"),(26,"Estado","R")]:
        pdf.cell(w,6,t(txt),align=al)
    pdf.ln(6)
    for titulo, key, fmt, meta, dirn in SERIE_SPEC:
        vals = [d[key] for d in serie if d[key] is not None]
        ap = vals[0] if vals else None; ci = vals[-1] if vals else None
        pr = (sum(vals)/len(vals)) if vals else None
        if meta is None or ci is None:
            edo = "-"
        else:
            signo = -1 if dirn=="menor" else 1
            edo = "Cumple" if (ci-meta)*signo >= 0 else "Falta"
        pdf.set_x(Mg); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","",8)
        pdf.cell(60,5.5,t(titulo),align="L")
        pdf.cell(26,5.5,t(_fserie(ap,fmt)),align="R")
        pdf.cell(26,5.5,t(_fserie(ci,fmt)),align="R")
        pdf.cell(26,5.5,t(_fserie(pr,fmt)),align="R")
        pdf.cell(22,5.5,t(_fserie(meta,fmt) if meta is not None else "-"),align="R")
        if edo=="Falta": pdf.set_text_color(192,57,43)
        elif edo=="Cumple": pdf.set_text_color(39,174,96)
        else: pdf.set_text_color(127,140,141)
        pdf.cell(26,5.5,t(edo),align="R"); pdf.ln(5.5)
    pdf.ln(3); pdf.set_text_color(127,140,141); pdf.set_font("Helvetica","",7.5)
    pdf.multi_cell(CW,3.6,t("Apertura = primer mes cargado; Cierre = ultimo mes; Promedio = media de los meses. "
                            "Metas: rentabilidad 10% y liquidez/endeudamiento por parametros de licitacion federal. "
                            "Para clientes en recontabilizacion (Fase 0) la serie es diagnostico, no linea base valida (R-MET-06)."))

    # ===== Estados financieros (cierre del ejercicio, comparativo vs anterior si existe) =====
    if ef_cierre is not None:
        ca = str(anio); cp = (str(anio-1) if ef_cierre_prev is not None else None)
        pdf.add_page()
        pdf.set_fill_color(28,45,58); pdf.rect(0,0,W,22,style="F")
        pdf.set_text_color(255,255,255); pdf.set_xy(Mg,5); pdf.set_font("Helvetica","B",14)
        pdf.cell(CW,7,t("Estados financieros - cierre " + str(anio)))
        pdf.set_xy(Mg,13); pdf.set_font("Helvetica","",9)
        pdf.cell(CW,5,t("Acumulado del ejercicio. Balance en posicion al cierre."))
        pdf.set_y(27)
        _pdf_tabla_estado(pdf, t, Mg, CW, "Estado de resultados", er_rows_cmp(ef_cierre, ef_cierre_prev), ca, cp)
        _pdf_tabla_estado(pdf, t, Mg, CW, "Balance general", bg_rows_cmp(ef_cierre, ef_cierre_prev), ca, cp)
        _pdf_tabla_estado(pdf, t, Mg, CW, "Estado de flujo de efectivo", efe_rows_cmp(ef_cierre, ef_cierre_prev), ca, cp)
        ok_efe, causa_efe, _acc = _efe_diag(ef_cierre["efe"])
        pdf.set_text_color((39,174,96) if ok_efe else (192,57,43)); pdf.set_font("Helvetica","",7.5)
        pdf.multi_cell(CW,3.6,t("Flujo cuadrado al cierre." if ok_efe else ("Flujo no cuadra: " + causa_efe)))

    # ===== Razones financieras (cierre del ejercicio) =====
    if rat_now:
        pdf.add_page()
        pdf.set_fill_color(28,45,58); pdf.rect(0,0,W,22,style="F")
        pdf.set_text_color(255,255,255); pdf.set_xy(Mg,5); pdf.set_font("Helvetica","B",14)
        pdf.cell(CW,7,t("Razones financieras - cierre " + str(anio)))
        pdf.set_xy(Mg,13); pdf.set_font("Helvetica","",9)
        pdf.cell(CW,5,t("Acumulado del ejercicio a la fecha de cierre, sin anualizar."))
        pdf.set_y(27)
        _pdf_tabla_ratios(pdf, t, Mg, CW, rat_now, rat_prev, str(anio), (str(anio-1) if rat_prev else None))
    # ===== Semáforos fiscales (cierre del ejercicio) =====
    if fiscal_sem:
        _cmap = {"rojo": (192,57,43), "amarillo": (241,196,15), "verde": (39,174,96),
                 "neutral": (149,165,166), "gris": (149,165,166)}
        pdf.add_page()
        pdf.set_fill_color(28,45,58); pdf.rect(0,0,W,22,style="F")
        pdf.set_text_color(255,255,255); pdf.set_xy(Mg,5); pdf.set_font("Helvetica","B",14)
        pdf.cell(CW,7,t("Semaforos fiscales - cierre " + str(anio)))
        pdf.set_xy(Mg,13); pdf.set_font("Helvetica","",9)
        pdf.cell(CW,5,t("Diagnostico fiscal interno. C2 desde contabilidad; C3/C4 con captura. C1 (69-B) en stand-by."))
        pdf.set_y(30)
        for luz, titulo, valor, msg, accion in fiscal_sem:
            col = _cmap.get(luz, (149,165,166))
            yb = pdf.get_y()
            pdf.set_draw_color(220,223,227); pdf.set_fill_color(248,249,250); pdf.rect(Mg, yb, CW, 22, style="DF")
            pdf.set_fill_color(*col); pdf.rect(Mg, yb, 3, 22, style="F")
            pdf.set_xy(Mg+7, yb+3); pdf.set_text_color(44,62,80); pdf.set_font("Helvetica","B",11)
            pdf.cell(CW-60, 6, t(titulo))
            pdf.set_xy(Mg+7+CW-60, yb+3); pdf.set_text_color(*col); pdf.set_font("Helvetica","B",13)
            pdf.cell(45, 6, t(valor), align="R")
            pdf.set_xy(Mg+7, yb+10); pdf.set_text_color(80,80,80); pdf.set_font("Helvetica","",9)
            pdf.multi_cell(CW-12, 4.2, t(msg))
            if accion:
                pdf.set_xy(Mg+7, yb+16.5); pdf.set_text_color(*col); pdf.set_font("Helvetica","I",8.5)
                pdf.multi_cell(CW-12, 4, t("Accion facturable: " + accion))
            pdf.set_y(yb+25)
    return bytes(pdf.output())


def _metas_cliente_rows(ind, metas):
    """Filas de metas en lenguaje de dueno: (indicador, hoy, meta, estado, ok)."""
    metas = metas or {}; rows = []
    p = ind.get("pretax_pct")
    if p is not None:
        m = meta_de("pretax_pct", metas); _, edo = evaluar_meta(p, m)
        mv = m["valor_meta"] if m else 10.0
        rows.append(("Rentabilidad (antes de impuestos)", "{:.1f}%".format(p), "{:.0f}%".format(mv),
                     ("Cumple" if edo == "cumple" else "Falta"), edo == "cumple"))
    ef = ind.get("_ef")
    if ef is not None:
        s = _stocks_de(ef)
        if s.get("pas_circ"):
            acpc = s["act_circ"] / s["pas_circ"]; ok = acpc >= LICIT_AC_PC
            rows.append(("Liquidez (activo / pasivo circulante)", "{:.2f}x".format(acpc),
                         "{:.2f}x".format(LICIT_AC_PC), ("Cumple" if ok else "Falta"), ok))
        if s.get("activo"):
            ptat = s["pasivo"] / s["activo"]; ok = ptat <= LICIT_PT_AT
            rows.append(("Endeudamiento (pasivo / activo)", "{:.1f}%".format(ptat*100),
                         "<={:.0f}%".format(LICIT_PT_AT*100), ("Cumple" if ok else "Falta"), ok))
    return rows

def pdf_reporte_cliente(nombre, periodo, ind, lectura, ef_a=None, ef_p=None, per_p=None, rat=None,
                        numero_mes=None, acciones=None, valor_generado=None, metas=None, poder_uno=None,
                        branding=None):
    """Reporte CFO mensual del cliente. Base YTD (acumulado). Caja al centro (regla de hierro)."""
    from fpdf import FPDF
    class _PDFCli(FPDF):
        def footer(self):
            self.set_y(-12); self.set_font("Helvetica", "", 7.5); self.set_text_color(150, 150, 150)
            self.set_x(self.l_margin)
            self.cell(0, 5, "Vastion Accounting  |  Confidencial  |  Pagina " + str(self.page_no()), align="C")
    def t(s):
        return (str(s).replace("\u2014", "-").replace("\u2013", "-").replace("\u2212", "-")
                .replace("\u0394", "Var.").encode("latin-1", "replace").decode("latin-1"))
    def m(v):
        return money(v) if v is not None else "s/d"
    W, Mg = 210, 16; CW = W - 2 * Mg
    pdf = _PDFCli(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=16)
    import io as _io
    _br = branding or {}
    _PRIM = _hex_rgb(_br.get("color_primario"), (28, 45, 58))
    _ACC = _hex_rgb(_br.get("color_acento"), (46, 134, 171))
    _GRY = (120, 120, 120); _INK = (40, 40, 40)
    _logo = _br.get("logo")

    def _hcli(titulo, sub=None, big=False):
        """Cabecera minimalista: logo + titulo en primario + linea de acento. Sin barra rellena."""
        if _logo:
            try:
                pdf.image(_io.BytesIO(_logo), x=W - Mg - 30, y=10, h=13)
            except Exception:
                pass
        pdf.set_xy(Mg, 14); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 19 if big else 16)
        pdf.cell(CW - 34, 9, t(titulo))
        _yl = 25
        if sub:
            pdf.set_xy(Mg, 24); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 11); pdf.cell(CW, 6, t(sub))
            _yl = 33
        pdf.set_draw_color(*_ACC); pdf.set_line_width(0.6); pdf.line(Mg, _yl, W - Mg, _yl); pdf.set_line_width(0.2)
        return _yl


    # ===================== HOJA 1: el mes en una hoja =====================
    pdf.add_page()
    _hcli("Reporte CFO Mensual", nombre + "   ·   " + periodo, big=True)

    pdf.set_xy(Mg, 42); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 11); pdf.cell(CW, 6, t("Tu caja al cierre del mes"))
    pdf.set_xy(Mg, 49); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 32); pdf.cell(CW, 14, t(m(ind.get("caja"))))
    cv = ind.get("caja_var")
    pdf.set_xy(Mg, 65); pdf.set_font("Helvetica", "", 11)
    if cv is not None:
        if cv < 0: pdf.set_text_color(192, 57, 43); _tx = "Bajo " + m(abs(cv)) + " respecto al mes anterior"
        else:      pdf.set_text_color(39, 174, 96); _tx = "Subio " + m(abs(cv)) + " respecto al mes anterior"
        pdf.cell(CW, 6, t(_tx))
    else:
        pdf.set_text_color(*_GRY); pdf.cell(CW, 6, t("Sin mes anterior para comparar"))

    yNum = 75
    if numero_mes and numero_mes.strip():
        pdf.set_draw_color(*_ACC); pdf.set_line_width(0.8); pdf.line(Mg, yNum, Mg, yNum + 13); pdf.set_line_width(0.2)
        pdf.set_xy(Mg + 5, yNum); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 8.5)
        pdf.cell(CW - 8, 4, t("El numero mas importante del mes"))
        pdf.set_xy(Mg + 5, yNum + 5); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 11)
        pdf.multi_cell(CW - 8, 5, t(numero_mes.strip()))
        y0 = yNum + 20
    else:
        y0 = yNum

    p = ind.get("pretax_pct")
    if   p is None: c_rent, n_rent = (149, 165, 166), "Sin dato"
    elif p >= 10:   c_rent, n_rent = (39, 174, 96),  "Sano"
    elif p >= 5:    c_rent, n_rent = (241, 196, 15), "Minimo"
    else:           c_rent, n_rent = (192, 57, 43),  "Peligro"
    v_rent = (str(p) + "%") if p is not None else "s/d"
    mb = ind.get("mb_pct"); v_mb = (str(mb) + "%") if mb is not None else "s/d"
    uai = ind.get("uai"); cl = ind.get("cash_lag")
    if   cl is None or uai is None: c_cl, n_cl = (149, 165, 166), "Sin dato"
    elif uai <= 0:                  c_cl, n_cl = (149, 165, 166), "Hay perdida; la prioridad es la rentabilidad"
    elif cl > 0:                    c_cl, n_cl = (192, 57, 43),  "La utilidad no llego a caja"
    else:                           c_cl, n_cl = (39, 174, 96),  "La caja siguio a la utilidad"

    def card(x, y, w, titulo, valor, color, nota, neutro=False):
        # Minimalista: solo linea superior (semaforo) + texto. Sin relleno ni borde.
        _ln = (205, 208, 212) if neutro else color
        pdf.set_draw_color(*_ln); pdf.set_line_width(0.8); pdf.line(x, y, x + w, y); pdf.set_line_width(0.2)
        pdf.set_xy(x, y + 3.5);  pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 8.5); pdf.multi_cell(w, 4, t(titulo))
        pdf.set_xy(x, y + 12); pdf.set_text_color(*_INK); pdf.set_font("Helvetica", "B", 16); pdf.cell(w, 8, t(valor))
        pdf.set_xy(x, y + 21); pdf.set_text_color(*(_GRY if neutro else color)); pdf.set_font("Helvetica", "", 8); pdf.multi_cell(w, 4, t(nota))

    gap = 5; cw = (CW - 2 * gap) / 3
    card(Mg,                y0, cw, "Rentabilidad del ano (acumulada)", v_rent, c_rent, n_rent)
    card(Mg + cw + gap,     y0, cw, "Margen de tu operacion",           v_mb,   _PRIM,  "Acumulado del ano", neutro=True)
    card(Mg + 2*(cw + gap), y0, cw, "Brecha utilidad vs. caja",         m(cl),  c_cl,   n_cl)

    yN = y0 + 32 + 6
    pdf.set_xy(Mg, yN); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 12); pdf.cell(CW, 7, t("Lo que esto significa"))
    pdf.set_draw_color(*_ACC); pdf.set_line_width(0.8); pdf.line(Mg, yN + 9, Mg, 274); pdf.set_line_width(0.2)
    pdf.set_xy(Mg + 5, yN + 10); pdf.set_font("Helvetica", "", 10.5); pdf.set_text_color(*_INK)
    _txt = lectura.strip() if (lectura and lectura.strip()) else "(La lectura del mes la escribe el CFO antes de enviar el reporte al cliente.)"
    pdf.multi_cell(CW - 8, 5.5, t(_txt))

    # ===================== HOJA 2: que hacer =====================
    pdf.add_page()
    _hcli("Que hacer este mes")
    pdf.set_y(32)
    pdf.set_x(Mg); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 12); pdf.cell(CW, 7, t("Las 3 acciones del mes")); pdf.ln(9)
    _acc = [a.strip() for a in (acciones or "").split("\n") if a.strip()][:3]
    if not _acc:
        pdf.set_x(Mg); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(CW, 5.5, t("(El CFO escribe aqui las 3 acciones concretas del mes, una por linea.)")); pdf.ln(2)
    else:
        for i, a in enumerate(_acc, 1):
            yb = pdf.get_y()
            pdf.set_fill_color(*_PRIM); pdf.set_text_color(255, 255, 255); pdf.set_font("Helvetica", "B", 11)
            pdf.set_xy(Mg, yb); pdf.cell(8, 8, t(str(i)), align="C", fill=True)
            pdf.set_xy(Mg + 12, yb); pdf.set_text_color(*_INK); pdf.set_font("Helvetica", "", 10.5)
            pdf.multi_cell(CW - 12, 5.5, t(a)); pdf.ln(3)
    pdf.ln(2); pdf.set_x(Mg); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 12); pdf.cell(CW, 7, t("Tus metas")); pdf.ln(9)
    pdf.set_x(Mg); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "B", 8)
    pdf.cell(82, 6, t("Indicador")); pdf.cell(32, 6, t("Hoy"), align="R"); pdf.cell(32, 6, t("Meta"), align="R"); pdf.cell(CW - 146, 6, t("Estado"), align="R"); pdf.ln(6)
    _fm = _metas_cliente_rows(ind, metas)
    if not _fm:
        pdf.set_x(Mg); pdf.set_text_color(127, 140, 141); pdf.set_font("Helvetica", "", 9); pdf.cell(CW, 6, t("Sin metas disponibles para este periodo.")); pdf.ln(6)
    for lbl, hoy, meta_v, edo, ok in _fm:
        pdf.set_x(Mg); pdf.set_text_color(44, 62, 80); pdf.set_font("Helvetica", "", 9)
        pdf.cell(82, 6, t(lbl)); pdf.cell(32, 6, t(hoy), align="R"); pdf.cell(32, 6, t(meta_v), align="R")
        if ok is None: pdf.set_text_color(127, 140, 141)
        elif ok:       pdf.set_text_color(39, 174, 96)
        else:          pdf.set_text_color(192, 57, 43)
        pdf.cell(CW - 146, 6, t(edo), align="R"); pdf.ln(6)

    # ===================== HOJA: las palancas (Poder del Uno) =====================
    if poder_uno:
        pu = poder_uno
        _tlbl = {"mensual": "del mes", "ytd": "acumulado del ano", "anualizado": "anualizado"}.get(pu.get("temporalidad"), "")
        def _sgn(v):
            return ("+" if v >= 0 else "-") + money(abs(v))
        pdf.add_page()
        _hcli("Las palancas de tu negocio")
        pdf.set_y(30)
        pdf.set_x(Mg); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(CW, 4.5, t("Cuanto cambia tu utilidad o tu caja si mueves cada palanca (base: " + _tlbl + ")."))
        pdf.ln(3)
        pdf.set_x(Mg); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "B", 8)
        pdf.cell(72, 6, t("Palanca")); pdf.cell(34, 6, t("Movimiento"), align="R"); pdf.cell(42, 6, t("Efecto"), align="R"); pdf.cell(CW - 148, 6, t("Sobre"), align="R"); pdf.ln(6)
        for nom, mov, imp, k in pu["filas"]:
            pdf.set_x(Mg); pdf.set_text_color(*_INK); pdf.set_font("Helvetica", "", 9)
            pdf.cell(72, 6, t(nom)); pdf.cell(34, 6, t(mov), align="R")
            if imp > 0: pdf.set_text_color(39, 174, 96)
            elif imp < 0: pdf.set_text_color(192, 57, 43)
            else: pdf.set_text_color(*_GRY)
            pdf.cell(42, 6, t(_sgn(imp)), align="R")
            pdf.set_text_color(*_GRY); pdf.cell(CW - 148, 6, t(k), align="R"); pdf.ln(6)
        pdf.ln(3)
        if pu.get("trampa"):
            yb = pdf.get_y()
            pdf.set_draw_color(192, 57, 43); pdf.set_line_width(0.8); pdf.line(Mg, yb, Mg, yb + 13); pdf.set_line_width(0.2)
            pdf.set_xy(Mg + 4, yb); pdf.set_text_color(192, 57, 43); pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(CW - 6, 4.5, t("Cuidado: con tu margen actual, vender mas volumen REDUCE la utilidad. "
                                          "Primero se arregla el margen, despues se busca vender mas."))
            pdf.set_y(yb + 16)
        pdf.set_x(Mg); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 11); pdf.cell(CW, 7, t("Si mueves todas las palancas juntas:")); pdf.ln(9)
        for _lbl, _v in [("Cambio en tu utilidad", pu.get("du", 0)), ("Cambio en tu caja", pu.get("dc", 0))]:
            pdf.set_x(Mg); pdf.set_text_color(*_INK); pdf.set_font("Helvetica", "", 10.5); pdf.cell(72, 8, t(_lbl))
            if _v > 0: pdf.set_text_color(39, 174, 96)
            elif _v < 0: pdf.set_text_color(192, 57, 43)
            else: pdf.set_text_color(*_GRY)
            pdf.set_font("Helvetica", "B", 13); pdf.cell(CW - 72, 8, t(_sgn(_v))); pdf.ln(9)

    # ===================== HOJA 3: valor generado =====================
    pdf.add_page()
    _hcli("Valor generado")
    pdf.set_y(34)
    if valor_generado and valor_generado.strip():
        pdf.set_x(Mg); pdf.set_text_color(*_INK); pdf.set_font("Helvetica", "", 11)
        pdf.multi_cell(CW, 6, t(valor_generado.strip()))
    else:
        pdf.set_x(Mg); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 10.5)
        pdf.multi_cell(CW, 5.5, t("(El CFO documenta aqui el valor generado por Vastion en el mes, en pesos: impuestos "
                                  "ahorrados, IVA recuperado, contingencias evitadas, ahorro financiero. Incluir el "
                                  "acumulado: 'Desde que Vastion es tu CFO hemos generado $X en valor documentado.')"))

    # ========================= ANEXO TÉCNICO =========================
    def _fila(label, sa, sp, sv, dual, bold=False, header=False):
        rh = 5.2
        if pdf.get_y() + rh > 281:
            pdf.add_page()
        lab = label if len(str(label)) <= 50 else (str(label)[:48] + "..")
        pdf.set_x(Mg)
        if header:
            pdf.set_fill_color(*_PRIM); pdf.set_text_color(255, 255, 255); pdf.set_font("Helvetica", "B", 9)
            pdf.cell(CW, 6, t(" " + lab), fill=True); pdf.ln(7); return
        if sa is None and sp is None and sv is None:
            pdf.set_fill_color(238, 240, 242); pdf.set_text_color(80, 90, 100); pdf.set_font("Helvetica", "B", 8)
            pdf.cell(CW, rh, t(" " + lab), fill=True); pdf.ln(rh); return
        pdf.set_text_color(44, 62, 80); pdf.set_font("Helvetica", "B" if bold else "", 8.5)
        if dual:
            pdf.cell(86, rh, t(lab)); pdf.cell(32, rh, t(sa), align="R")
            pdf.cell(30, rh, t(sp if sp is not None else ""), align="R"); pdf.cell(30, rh, t(sv if sv is not None else ""), align="R")
        else:
            pdf.cell(118, rh, t(lab)); pdf.cell(60, rh, t(sa), align="R")
        pdf.ln(rh)

    def _colhead(dual, ca, cpx):
        pdf.set_x(Mg); pdf.set_text_color(127, 140, 141); pdf.set_font("Helvetica", "B", 7.5)
        if dual:
            pdf.cell(86, 5, t("Concepto")); pdf.cell(32, 5, t(ca), align="R")
            pdf.cell(30, 5, t(cpx if cpx else ""), align="R"); pdf.cell(30, 5, t("Variacion"), align="R")
        else:
            pdf.cell(118, 5, t("Concepto")); pdf.cell(60, 5, t(ca), align="R")
        pdf.ln(5)

    def _estado(titulo, rows, dual, ca, cpx):
        if pdf.get_y() + 24 > 281: pdf.add_page()
        _fila(titulo, None, None, None, dual, header=True)
        _colhead(dual, ca, cpx)
        for label, va, vp, bold in rows:
            if va is None and vp is None:
                _fila(label, None, None, None, dual)
            else:
                sa = _fmt(va)
                sp = _fmt(vp) if dual else None
                sv = _fmt((va or 0) - (vp or 0)) if dual else None
                _fila(label, sa, sp, sv, dual, bold=bold)
        pdf.ln(3)

    def _banner_anexo(titulo, nota):
        pdf.add_page()
        pdf.set_xy(Mg, 14); pdf.set_text_color(*_PRIM); pdf.set_font("Helvetica", "B", 13)
        pdf.cell(CW, 8, t(titulo))
        pdf.set_draw_color(*_ACC); pdf.set_line_width(0.6); pdf.line(Mg, 24, W - Mg, 24); pdf.set_line_width(0.2)
        pdf.set_xy(Mg, 27); pdf.set_text_color(*_GRY); pdf.set_font("Helvetica", "", 7.5)
        pdf.multi_cell(CW, 4, t(nota)); pdf.ln(2)

    cp_lbl = (per_p[:7] if per_p else None)
    if ef_a is not None:
        _banner_anexo("Anexo - Estados financieros  ·  " + periodo,
                      "ER y EFE en acumulado del ejercicio (YTD); Balance en posicion del mes. "
                      "Para clientes en recontabilizacion son diagnostico interno hasta cerrar Fase 0.")
        ok_efe, causa_efe, acc_efe = _efe_diag(ef_a["efe"])
        if not ok_efe:
            yb = pdf.get_y()
            pdf.set_fill_color(192, 57, 43); pdf.rect(Mg, yb, CW, 8, style="F")
            pdf.set_xy(Mg + 2, yb + 1.5); pdf.set_text_color(255, 255, 255); pdf.set_font("Helvetica", "B", 9)
            pdf.cell(CW - 4, 5, t("ESTADOS RETENIDOS - R-EFE-01: el flujo de efectivo no cuadra"))
            pdf.set_xy(Mg, yb + 9); pdf.set_text_color(192, 57, 43); pdf.set_font("Helvetica", "", 7.5)
            pdf.multi_cell(CW, 3.6, t(causa_efe + "  Accion: " + acc_efe))
            pdf.ln(2); pdf.set_text_color(0, 0, 0)
        dual = ef_p is not None
        _estado("Estado de Resultados", er_rows_cmp(ef_a, ef_p), dual, periodo, cp_lbl)
        _estado("Balance General", bg_rows_cmp(ef_a, ef_p), dual, periodo, cp_lbl)
        _estado("Estado de Flujo de Efectivo", efe_rows_cmp(ef_a, ef_p), dual, periodo, cp_lbl)

    if rat is not None and rat[0]:
        a_rows, p_rows, rper = rat
        _banner_anexo("Anexo - Razones financieras  ·  " + periodo,
                      "Acumulado del anio a la fecha (YTD), sin anualizar. Los dias reflejan los dias transcurridos del anio.")
        pdual = bool(p_rows)
        pmap = {(s, l): (v, f) for s, l, v, f in p_rows} if p_rows else {}
        cpr = (rper[:7] if rper else None)
        sec_now = None
        for s, l, v, f in a_rows:
            if s != sec_now:
                if pdf.get_y() + 14 > 281: pdf.add_page()
                _fila(s, None, None, None, pdual, header=True)
                _colhead(pdual, periodo, cpr)
                sec_now = s
            vp = pmap.get((s, l), (None, f))[0]
            sa = _fr(v, f); sp = _fr(vp, f) if pdual else None
            sv = _fvar(v, vp, f) if pdual else None
            _fila(l, sa, sp, sv, pdual)

    return bytes(pdf.output())


def tendencia(cli):
    conn = get_conn(); cur = conn.cursor(); out = []
    try:
        cur.execute("SELECT periodo, archivo_vigente, estado FROM periodo_estado WHERE cliente_id=%s ORDER BY periodo", (cli,))
        for per, bal, estado in cur.fetchall():
            if bal is None: continue
            cur.execute("SELECT pl.pct FROM fn_pl_gestion(%s) pl WHERE pl.concepto='= UTILIDAD ANTES DE IMPUESTOS'", (bal,))
            r = cur.fetchone(); pretax = float(r[0]) if r and r[0] is not None else None
            cur.execute("SELECT concepto,valor FROM fn_cash_lag(%s)", (bal,))
            cl = {c:(float(v) if v is not None else None) for c,v in cur.fetchall()}
            caja = cl.get('Caja fin de mes')
            cash_lag = next((v for k,v in cl.items() if k.startswith('Cash Lag')), None)
            cur.execute("SELECT 1 FROM fn_validar_madurez(%s) WHERE paso=false LIMIT 1", (bal,)); adv_m = cur.fetchone() is not None
            cur.execute("SELECT 1 FROM fn_validar_periodo(%s) WHERE paso=false LIMIT 1", (bal,)); adv_i = cur.fetchone() is not None
            out.append(dict(periodo=str(per)[:7], pretax=pretax, caja=caja, cash_lag=cash_lag, flag=(adv_m or adv_i)))
        return out
    finally:
        cur.close(); conn.close()


def datos_reporte_cliente(cli, periodo):
    """Datos del reporte del cliente, base YTD (acumulado). Caja = agrupadores 101/102/103 (igual que el flujo)."""
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT archivo_vigente FROM periodo_estado WHERE cliente_id=%s AND periodo=%s", (cli, periodo))
        r = cur.fetchone()
        if not r or not r[0]: return None
        bal = r[0]
        cur.execute("""SELECT archivo_vigente FROM periodo_estado
                       WHERE cliente_id=%s AND periodo<%s AND archivo_vigente IS NOT NULL
                       ORDER BY periodo DESC LIMIT 1""", (cli, periodo))
        rp = cur.fetchone(); bal_prev = rp[0] if rp else None
    finally:
        cur.close(); conn.close()
    ef = estados_financieros(bal); ytd = ef["ytd"]; ing = ytd["ing"]
    ind = {}
    ind["pretax_pct"] = round(ytd["uai"] / ing * 100, 1) if ing else None
    ind["mb_pct"]     = round(ytd["ub"] / ing * 100, 1) if ing else None
    ind["uai"]        = ytd["uai"]
    ind["caja"]       = ef["efe"]["efec_fin"]
    ind["cash_lag"]   = ytd["uai"] - (ef["efe"]["efec_fin"] - ef["efe"]["efec_ini"])
    caja_prev = estados_financieros(bal_prev)["efe"]["efec_fin"] if bal_prev else None
    ind["caja_prev"]  = caja_prev
    ind["caja_var"]   = (ind["caja"] - caja_prev) if caja_prev is not None else None
    ind["_ef"]        = ef
    return ind

def periodos_cargados(cli):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT periodo FROM periodo_estado WHERE cliente_id=%s AND archivo_vigente IS NOT NULL
                       ORDER BY periodo DESC""", (cli,))
        return [str(r[0]) for r in cur.fetchall()]
    finally:
        cur.close(); conn.close()


def periodos_cfdi_emitido(cli):
    """Meses con CFDI emitidos cargados, leidos de raw_cfdi. Para el selector de la mezcla:
       refleja exactamente lo que hay, sin depender de que exista balanza de ese mes."""
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT DISTINCT periodo FROM raw_cfdi
                       WHERE cliente_id=%s AND direccion='EMITIDO' ORDER BY periodo DESC""", (cli,))
        return [str(r[0]) for r in cur.fetchall()]
    finally:
        cur.close(); conn.close()


def top_clientes(cli, periodo, n=10):
    """Top-N clientes por ingreso neto emitido (sin IVA, vigente, neto de notas de credito).
       Llave = RFC (Codigo Cliente viene vacio). Devuelve 'mes' (periodo exacto) y
       'acumulado' (YTD del ejercicio: del 1-ene del anio del periodo al periodo, reinicia
       en enero igual que ER/EFE). Trae total, # clientes, lista top y concentracion. Alimenta Alerta #3."""
    conn = get_conn(); cur = conn.cursor(); out = {}
    try:
        for clave in ("mes", "acumulado"):
            if clave == "mes":
                cond, params = "periodo = %s", (cli, periodo)
            else:  # YTD del ejercicio (reinicia en enero, consistente con estados financieros)
                cond, params = ("periodo >= date_trunc('year', %s::date) AND periodo <= %s",
                                (cli, periodo, periodo))
            cur.execute(f"""
                WITH base AS (
                  SELECT contraparte_rfc AS rfc, contraparte_nom AS nom, uuid, tipo_cfdi,
                         (COALESCE(subtotal,0)-COALESCE(descuento,0)) AS neto
                  FROM raw_cfdi
                  WHERE cliente_id=%s AND direccion='EMITIDO'
                    AND lower(estatus_sat)='vigente'
                    AND lower(tipo_cfdi) IN ('ingreso','egreso')
                    AND {cond}
                ),
                por_cliente AS (
                  SELECT rfc, max(nom) AS cliente, count(DISTINCT uuid) AS facturas,
                         sum(CASE WHEN lower(tipo_cfdi)='egreso' THEN -neto ELSE neto END) AS ingreso_neto
                  FROM base GROUP BY rfc
                )
                SELECT rfc, cliente, ingreso_neto, facturas
                FROM por_cliente ORDER BY ingreso_neto DESC
            """, params)
            filas = cur.fetchall()
            total = sum(float(f[2] or 0) for f in filas)
            top = [dict(rfc=f[0], cliente=f[1], neto=float(f[2] or 0), facturas=f[3],
                        pct=(round(100*float(f[2] or 0)/total, 1) if total else None)) for f in filas[:n]]
            conc1 = round(100*top[0]["neto"]/total, 1) if top and total else None
            conc3 = round(100*sum(t["neto"] for t in top[:3])/total, 1) if top and total else None
            out[clave] = dict(total=total, clientes=len(filas), top=top, conc_top1=conc1, conc_top3=conc3)
        return out
    finally:
        cur.close(); conn.close()


SEG_SAT = {
    '10':'Material vivo vegetal y animal','11':'Material mineral y textil','12':'Productos quimicos',
    '13':'Resinas, caucho y espumas','14':'Papel y productos de papel','15':'Combustibles y lubricantes',
    '20':'Maquinaria de mineria y perforacion','21':'Maquinaria agricola','22':'Maquinaria de construccion',
    '23':'Maquinaria industrial','24':'Manejo y almacenaje de materiales','25':'Vehiculos y transporte',
    '26':'Generacion y distribucion de energia','27':'Herramientas y maquinaria general','30':'Materiales de construccion y estructuras',
    '31':'Componentes de manufactura','32':'Componentes y suministros electronicos','39':'Equipo electrico e iluminacion',
    '40':'Climatizacion y distribucion de fluidos','41':'Equipo de laboratorio y medicion','42':'Equipo y suministros medicos',
    '43':'Tecnologia de la informacion','44':'Equipo y suministros de oficina','45':'Equipo de impresion y fotografia',
    '46':'Equipo de defensa y seguridad','47':'Equipo de limpieza','48':'Equipo de hosteleria y comercio',
    '49':'Equipo deportivo y recreativo','50':'Alimentos y bebidas','51':'Productos farmaceuticos','52':'Articulos para el hogar',
    '53':'Ropa, calzado y equipaje','55':'Productos impresos y editoriales','56':'Mobiliario','60':'Instrumentos musicales y arte',
    '70':'Servicios agricolas, pesca y forestal','71':'Servicios de mineria, petroleo y gas',
    '72':'Construccion, instalaciones y mantenimiento','73':'Servicios de produccion industrial','76':'Servicios de limpieza industrial',
    '77':'Servicios medioambientales','78':'Transporte, almacenaje y correo','80':'Servicios de gestion y administrativos',
    '81':'Ingenieria, investigacion y tecnologia','82':'Servicios editoriales, diseno y artes graficas','83':'Servicios publicos (utilities)',
    '84':'Servicios financieros y de seguros','85':'Servicios de salud','86':'Servicios educativos',
    '90':'Viajes, alimentacion, hospedaje y entretenimiento','91':'Servicios personales y domesticos',
    '92':'Seguridad y orden publico','93':'Servicios politicos y de organizaciones','94':'Organizaciones y asociaciones',
    '95':'Terrenos, edificios y estructuras',
}


def mezcla_ingresos(cli, periodo):
    """Modelo de negocio: mezcla del ingreso emitido por segmento SAT (2 primeros digitos de raw_cfdi.clave_sat).
       Lee DIRECTO de la base, sin re-subir archivos. Ingreso neto = subtotal - descuento (sin IVA), solo Vigente,
       neto de notas de credito (egreso resta). Ventana YTD del ejercicio del mes seleccionado.
       clave_sat se captura al cargar el CFDI; los comprobantes cargados antes de esa captura salen sin clave."""
    y = int(str(periodo)[:4]); ene1 = date(y, 1, 1)
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT substr(clave_sat, 1, 2) AS seg,
                   count(DISTINCT uuid) AS fac,
                   sum(CASE WHEN lower(tipo_cfdi)='egreso' THEN -1 ELSE 1 END
                       * (COALESCE(subtotal,0) - COALESCE(descuento,0))) AS neto
            FROM raw_cfdi
            WHERE cliente_id=%s AND direccion='EMITIDO'
              AND lower(estatus_sat)='vigente' AND lower(tipo_cfdi) IN ('ingreso','egreso')
              AND periodo >= %s AND periodo <= %s
            GROUP BY substr(clave_sat, 1, 2)
            ORDER BY neto DESC
        """, (cli, ene1, periodo))
        filas = cur.fetchall()
        cur.execute("""SELECT count(DISTINCT uuid) FILTER (WHERE clave_sat IS NOT NULL),
                              count(DISTINCT uuid)
                       FROM raw_cfdi
                       WHERE cliente_id=%s AND direccion='EMITIDO'
                         AND lower(estatus_sat)='vigente' AND lower(tipo_cfdi) IN ('ingreso','egreso')
                         AND periodo >= %s AND periodo <= %s""", (cli, ene1, periodo))
        con_clave, total_cfdi = cur.fetchone()
    finally:
        cur.close(); conn.close()
    con_clave = int(con_clave or 0); total_cfdi = int(total_cfdi or 0)
    tot = sum(float(f[2] or 0) for f in filas)
    rows = []; sin_clave = 0.0
    for seg, fac, neto in filas:
        neto = float(neto or 0)
        if abs(neto) < 1: continue
        if not seg:                                    # comprobantes sin clave_sat (cargados antes de la captura)
            sin_clave += neto; continue
        rows.append(dict(seg=seg, etiqueta=SEG_SAT.get(seg, "Segmento " + seg),
                         neto=neto, facturas=fac,
                         pct=(round(100 * neto / tot, 1) if tot else None)))
    pct_sin = round(100 * sin_clave / tot, 1) if tot else 0.0
    return dict(rows=rows, total=tot, ejercicio=y,
                dominante=(rows[0] if rows else None), pct_sin_clave=pct_sin,
                con_clave=con_clave, total_cfdi=total_cfdi,
                cob_clave=(round(100 * con_clave / total_cfdi, 1) if total_cfdi else 0.0))


def cartera_clientes(cli, periodo):
    """Cartera por cliente desde la BALANZA (cuentas hoja agrupador 105): saldo autoritativo, reconcilia por
       construccion. DSO por cliente = saldo / facturado 12 meses moviles x 365 (Opcion 2, rotacion). Si no hay
       facturacion en 12m -> 'CONGELADA' (saldo parado, prioridad de cobranza). Facturado/ultimo mov del auxiliar."""
    y = int(str(periodo)[:4]); m = int(str(periodo)[5:7])
    nxt = date(y+1, 1, 1) if m == 12 else date(y, m+1, 1)
    corte = nxt - timedelta(days=1)                      # fin del mes seleccionado, para la antiguedad
    _mo = (y*12 + (m-1)) - 11
    t12_ini = date(_mo // 12, _mo % 12 + 1, 1)   # inicio de la ventana de 12 meses moviles
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT id FROM origen_archivo WHERE cliente_id=%s AND tipo='CATALOGO' AND periodo<=%s
                       ORDER BY periodo DESC, version DESC LIMIT 1""", (cli, periodo))
        r = cur.fetchone(); cat_id = r[0] if r else None
        cur.execute("""
            SELECT b.num_cuenta, c.descripcion, b.saldo_final
            FROM insumos_balanza b
            LEFT JOIN raw_catalogo c ON c.archivo_id=%s AND c.num_cuenta=b.num_cuenta
            WHERE b.cliente_id=%s AND b.periodo=%s AND b.es_hoja AND b.cod_agrupador LIKE '105%%'
            ORDER BY b.saldo_final DESC
        """, (cat_id, cli, periodo))
        filas = cur.fetchall()
        cur.execute("""
            SELECT subcuenta, max(tercero) AS tercero,
                   max(fecha) FILTER (WHERE NOT es_saldo_ini) AS ult_mov,
                   COALESCE(sum(debe) FILTER (WHERE tipo_poliza='Factura' AND fecha>=%s AND fecha<%s), 0) AS fac12
            FROM raw_aux_clientes WHERE cliente_id=%s GROUP BY subcuenta
        """, (t12_ini, nxt, cli))
        aux = {s: dict(ter=t, ult=u, fac12=float(f or 0)) for s, t, u, f in cur.fetchall()}
        # Estimacion para cuentas incobrables (agrupador SAT 108, acreedora, netea contra 105).
        # Si no esta agrupada en el catalogo del cliente, no se encuentra -> el tablero lo advierte, no lo esconde.
        cur.execute("""SELECT COALESCE(sum(saldo_final), 0)
                       FROM insumos_balanza
                       WHERE cliente_id=%s AND periodo=%s AND es_hoja AND cod_agrupador LIKE '108%%'""",
                    (cli, periodo))
        _est = cur.fetchone()
        estim = abs(float(_est[0] or 0)) if _est else 0.0
        estim_ok = estim > 1            # se encontro estimacion agrupada en 108
    finally:
        cur.close(); conn.close()
    out = []
    for num, desc, saldo in filas:
        saldo = float(saldo or 0)
        if abs(saldo) < 1: continue
        a = aux.get(num)
        nombre = desc or (a["ter"] if a else None) or num
        if a and a["fac12"] > 1:
            dso = round(saldo / a["fac12"] * 365)
        elif a:                       # hay historia en el auxiliar pero cero factura en 12m
            dso = "CONGELADA"
        else:                         # sin auxiliar para esta cuenta
            dso = None
        ult = a["ult"] if a and a.get("ult") else None
        antig = (corte - ult).days if ult else None      # dias desde el ultimo movimiento al cierre del mes
        out.append(dict(cuenta=num, cliente=nombre, saldo=saldo, dso=dso,
                        fac12=(a["fac12"] if a else 0.0),
                        ult_mov=(str(ult)[:10] if ult else None), antiguedad=antig))
    tot = sum(o["saldo"] for o in out)
    for o in out:
        o["pct"] = round(100 * o["saldo"] / tot, 1) if tot else None
    neto = tot - estim if estim_ok else None
    cobertura = round(100 * estim / tot, 1) if (estim_ok and tot) else None
    return dict(rows=out, total=tot, ejercicio=y,
                estim=estim, estim_ok=estim_ok, neto=neto, cobertura=cobertura)


def velocidad_cobro(cli, periodo):
    """Velocidad de cobro real (dias factura->cobro) desde el auxiliar de Clientes (raw_aux_clientes),
       fuente autoritativa que reconcilia con la balanza. NO usa la Fecha de pago del CFDI (no fiable:
       produce dias negativos y contradice la cartera). Empareja cada cobro (Movimiento Conciliado) con
       sus facturas mas viejas por FIFO; dias = cobro - factura, ponderado por monto. Ventana: cobros del
       ejercicio del mes seleccionado (ene1 a fin de mes); las facturas pueden ser de cualquier ejercicio
       (un cobro de 2026 puede pagar una factura de 2025). Ignora dias negativos (anticipos: cobro antes de
       factura). Los clientes congelados no aparecen aqui (no cobraron) -> viven en Cartera como CONGELADA."""
    y = int(str(periodo)[:4]); m = int(str(periodo)[5:7])
    ene1 = date(y, 1, 1)
    nxt = date(y+1, 1, 1) if m == 12 else date(y, m+1, 1)   # fin de mes exclusivo
    conn = get_conn(); cur = conn.cursor()
    try:
        # Facturas (cargos) de toda la historia, por subcuenta, para el pool de emparejamiento FIFO.
        cur.execute("""SELECT subcuenta, max(tercero), fecha, sum(debe)
                       FROM raw_aux_clientes
                       WHERE cliente_id=%s AND NOT es_saldo_ini AND tipo_poliza='Factura' AND debe>0
                       GROUP BY subcuenta, fecha ORDER BY subcuenta, fecha""", (cli,))
        fac = {}; nombre = {}
        for sub, ter, fch, debe in cur.fetchall():
            fac.setdefault(sub, []).append([fch, float(debe or 0)])
            if ter: nombre[sub] = ter
        # Cobros conciliados dentro del ejercicio del mes seleccionado.
        cur.execute("""SELECT subcuenta, fecha, sum(haber)
                       FROM raw_aux_clientes
                       WHERE cliente_id=%s AND NOT es_saldo_ini AND tipo_poliza='Movimiento Conciliado'
                         AND haber>0 AND fecha>=%s AND fecha<%s
                       GROUP BY subcuenta, fecha ORDER BY subcuenta, fecha""", (cli, ene1, nxt))
        cob = {}
        for sub, fch, haber in cur.fetchall():
            cob.setdefault(sub, []).append([fch, float(haber or 0)])
    finally:
        cur.close(); conn.close()
    rows = []; tot_dm = 0.0; tot_m = 0.0
    for sub, cobros in cob.items():
        pool = [list(x) for x in sorted(fac.get(sub, []))]   # copia mutable, mas viejas primero
        i = 0; dm = 0.0; m_emp = 0.0
        for fch_c, amt in sorted(cobros):
            rem = amt
            while rem > 0.01 and i < len(pool):
                fch_f, disp = pool[i]
                take = min(rem, disp)
                dias = (fch_c - fch_f).days
                if dias >= 0:                      # ignora anticipos (cobro antes de la factura)
                    dm += dias * take; m_emp += take
                pool[i][1] = disp - take; rem -= take
                if pool[i][1] <= 0.01: i += 1
        if m_emp > 1:
            rows.append(dict(sub=sub, cliente=nombre.get(sub, sub),
                             dias=round(dm / m_emp), cobrado=m_emp))
            tot_dm += dm; tot_m += m_emp
    rows.sort(key=lambda x: -x["cobrado"])
    agregado = round(tot_dm / tot_m) if tot_m else None
    return dict(rows=rows, ejercicio=y, agregado=agregado, tot_cobrado=tot_m)


def fraccionamiento(cli, periodo, umbral=1_000_000):
    """Control 1 (riesgo 69-B / materialidad) - patron de fraccionamiento en CFDI EMITIDOS.
       Marca grupos de 2+ comprobantes vigentes de ingreso, mismo dia (fecha de emision) y mismo receptor (RFC),
       donde cada CFDI es < umbral pero la SUMA del dia >= umbral: pedazos sub-umbral que reconstruyen una
       operacion que habria cruzado el umbral. Es SEÑAL DE RIESGO, no veredicto: obra publica se factura por
       estimacion/partida y puede explicar el patron; el tablero marca para revisar, no acusa. Ventana: YTD del
       ejercicio del mes. Un total por UUID (renglon=1); la hora no se usa (el export no la trae, solo el dia)."""
    y = int(str(periodo)[:4]); ene1 = date(y, 1, 1)
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            WITH base AS (
                SELECT uuid, fecha_emision::date AS dia, contraparte_rfc AS rfc,
                       max(contraparte_nom) AS nom, max(total) AS total
                FROM raw_cfdi
                WHERE cliente_id=%s AND direccion='EMITIDO' AND renglon=1
                  AND lower(estatus_sat)='vigente' AND lower(tipo_cfdi)='ingreso'
                  AND periodo >= %s AND periodo <= %s
                GROUP BY uuid, fecha_emision::date, contraparte_rfc
            )
            SELECT dia, rfc, max(nom) AS nom, count(*) AS n, sum(total) AS suma
            FROM base
            WHERE total < %s
            GROUP BY dia, rfc
            HAVING count(*) >= 2 AND sum(total) >= %s
            ORDER BY sum(total) DESC
        """, (cli, ene1, periodo, umbral, umbral))
        filas = cur.fetchall()
    finally:
        cur.close(); conn.close()
    rows = [dict(dia=str(d), rfc=r, cliente=(n or r), n=int(k), suma=float(s or 0))
            for d, r, n, k, s in filas]
    return dict(rows=rows, ejercicio=y, umbral=umbral,
                n_grupos=len(rows), n_cfdi=sum(x["n"] for x in rows),
                suma_total=sum(x["suma"] for x in rows))


def edad_rfc(rfc):
    """Edad del contribuyente a partir de la estructura del RFC. Persona moral: 12 caracteres
       (3 letras + AAMMDD + 3 homoclave); fisica: 13 (4 letras + AAMMDD + 3). Devuelve (tipo, fecha, anios).
       Pivote de siglo: interpreta AA como 20AA salvo que caiga en el futuro, entonces 19AA. Los RFC genericos
       (XAXX010101000 publico, XEXX010101000 extranjero) no tienen edad util -> ('generico', None, None)."""
    r = (rfc or "").strip().upper()
    if r in ("XAXX010101000", "XEXX010101000"):
        return ("generico", None, None)
    if len(r) == 12:
        fecha, tipo = r[3:9], "moral"
    elif len(r) == 13:
        fecha, tipo = r[4:10], "fisica"
    else:
        return ("invalido", None, None)
    if not fecha.isdigit():
        return (tipo, None, None)
    yy, mm, dd = int(fecha[:2]), int(fecha[2:4]), int(fecha[4:6])
    if not (1 <= mm <= 12 and 1 <= dd <= 31):
        return (tipo, None, None)
    hoy = date.today()
    try:
        f = date(2000 + yy, mm, dd)
        if f > hoy:
            f = date(1900 + yy, mm, dd)
    except ValueError:
        return (tipo, None, None)
    return (tipo, f, round((hoy - f).days / 365.25, 1))


def periodos_cfdi_recibido(cli):
    """Meses con CFDI recibidos cargados, leidos de raw_cfdi. Para el selector del indicador de proveedores."""
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT DISTINCT periodo FROM raw_cfdi
                       WHERE cliente_id=%s AND direccion='RECIBIDO' ORDER BY periodo DESC""", (cli,))
        return [str(r[0]) for r in cur.fetchall()]
    finally:
        cur.close(); conn.close()


def proveedores_69b(cli, periodo, umbral=1_000_000, edad_max=3.0):
    """Indicador de proveedores riesgo 69-B (CFDI RECIBIDOS). Dos criterios:
       (A) Fraccionamiento: 2+ CFDI recibidos vigentes de ingreso el mismo dia del mismo proveedor (RFC),
           cada uno < umbral pero sumando >= umbral. Mismo patron que emitidos, ahora por proveedor.
       (B) Edad del proveedor desde el RFC: proveedores jovenes con gasto relevante = bandera EFOS clasica.
       Ambos son SEÑAL DE RIESGO, no veredicto. Ventana: YTD del ejercicio del mes. Un total por UUID (renglon=1)."""
    y = int(str(periodo)[:4]); ene1 = date(y, 1, 1)
    conn = get_conn(); cur = conn.cursor()
    try:
        # (A) Fraccionamiento por dia + proveedor
        cur.execute("""
            WITH base AS (
                SELECT uuid, fecha_emision::date AS dia, contraparte_rfc AS rfc,
                       max(contraparte_nom) AS nom, max(total) AS total
                FROM raw_cfdi
                WHERE cliente_id=%s AND direccion='RECIBIDO' AND renglon=1
                  AND lower(estatus_sat)='vigente' AND lower(tipo_cfdi)='ingreso'
                  AND periodo >= %s AND periodo <= %s
                GROUP BY uuid, fecha_emision::date, contraparte_rfc
            )
            SELECT dia, rfc, max(nom) AS nom, count(*) AS n, sum(total) AS suma
            FROM base WHERE total < %s
            GROUP BY dia, rfc
            HAVING count(*) >= 2 AND sum(total) >= %s
            ORDER BY sum(total) DESC
        """, (cli, ene1, periodo, umbral, umbral))
        frac = cur.fetchall()
        # (B) Gasto por proveedor (neto de notas de credito) para la vista de edad
        cur.execute("""
            SELECT contraparte_rfc, max(contraparte_nom),
                   sum(CASE WHEN lower(tipo_cfdi)='egreso' THEN -1 ELSE 1 END
                       * (COALESCE(subtotal,0)-COALESCE(descuento,0))) AS gasto,
                   count(DISTINCT uuid) AS ncfdi
            FROM raw_cfdi
            WHERE cliente_id=%s AND direccion='RECIBIDO'
              AND lower(estatus_sat)='vigente' AND lower(tipo_cfdi) IN ('ingreso','egreso')
              AND periodo >= %s AND periodo <= %s
            GROUP BY contraparte_rfc
        """, (cli, ene1, periodo))
        prov = cur.fetchall()
    finally:
        cur.close(); conn.close()
    frac_rows = []
    for d, rfc, nom, n, s in frac:
        tipo, fconst, anios = edad_rfc(rfc)
        frac_rows.append(dict(dia=str(d), rfc=rfc, proveedor=(nom or rfc), n=int(n),
                              suma=float(s or 0), tipo=tipo, edad=anios,
                              joven=(anios is not None and anios < edad_max)))
    edad_rows = []
    for rfc, nom, gasto, ncfdi in prov:
        tipo, fconst, anios = edad_rfc(rfc)
        if anios is None:
            continue
        edad_rows.append(dict(rfc=rfc, proveedor=(nom or rfc), gasto=float(gasto or 0),
                              ncfdi=int(ncfdi), tipo=tipo, edad=anios,
                              constitucion=str(fconst), joven=(anios < edad_max)))
    edad_rows.sort(key=lambda x: (x["edad"], -x["gasto"]))   # mas jovenes primero
    jovenes = [r for r in edad_rows if r["joven"] and abs(r["gasto"]) >= 1]
    return dict(ejercicio=y, umbral=umbral, edad_max=edad_max,
                frac_rows=frac_rows, jovenes=jovenes,
                n_prov=len([r for r in edad_rows if abs(r["gasto"]) >= 1]))


SAT_LBL = {
    '101':'Caja','102':'Bancos','103':'Inversiones','104':'Inversiones',
    '105':'Clientes','106':'Documentos por cobrar','107':'Deudores diversos',
    '108':'IVA acreditable pagado','109':'Anticipo a proveedores','110':'Otras cuentas por cobrar',
    '113':'Impuestos a favor','114':'Pagos provisionales','115':'Inventarios',
    '116':'Obra en proceso','117':'Obra en proceso','118':'IVA acreditable',
    '119':'IVA pendiente de acreditar','120':'Anticipo a proveedores','121':'Inventarios',
    '126':'Pagos anticipados','151':'Terrenos','152':'Edificios','153':'Maquinaria y equipo',
    '154':'Equipo de transporte','155':'Equipo de cómputo','156':'Mobiliario y equipo',
    '157':'Equipo','158':'Equipo','159':'(−) Depreciación acumulada','171':'(−) Depreciación acumulada',
    '172':'(−) Amortización acumulada','174':'Activos intangibles','180':'Activos diferidos','184':'Otros activos',
    '201':'Proveedores','202':'Créditos bancarios CP','203':'Cuentas por pagar','204':'Cuentas por pagar',
    '205':'Acreedores diversos','206':'Cuentas por pagar','207':'IVA trasladado cobrado','208':'Anticipo de clientes',
    '209':'IVA trasladado por cobrar','210':'Provisiones','211':'Provisiones de nómina','213':'Impuestos por pagar',
    '214':'Impuestos por pagar','215':'PTU por pagar','216':'Impuestos retenidos',
    '251':'Créditos bancarios LP','252':'Documentos por pagar LP','253':'Acreedores LP',
    '301':'Capital social y aportaciones','302':'Aportaciones','303':'Reservas',
    '304':'Resultados de ejercicios anteriores','305':'Resultado del ejercicio',
}
def _lbl(code): return SAT_LBL.get(code, "Agrupador " + (code or "?"))

def _fiscal_buckets(acc, saldo_g):
    """Identifica no deducible y accesorios DESDE contabilidad, por subcódigo agrupador SAT (YTD). R-FIS-04/08.
    Permanente = recargos (6xx.59) + multas/sanciones/actualización (6xx.84).
    Sin requisitos = 6xx.83 / 6xx.81. No deducible CUFIN = 612. ISR provisional pagado = agrupador 114."""
    def _ssf(x): return x['sf'] if x['nat'] == 'D' else -x['sf']
    def _suf(ag):
        p = (ag or '').split('.'); return p[1] if len(p) > 1 else ''
    s6 = [x for x in acc if (x['ag'] or '')[:1] == '6']
    recargos = sum(_ssf(x) for x in s6 if _suf(x['ag']) == '59')
    multas   = sum(_ssf(x) for x in s6 if _suf(x['ag']) == '84')
    sinreq   = sum(_ssf(x) for x in s6 if _suf(x['ag']) in ('83', '81'))
    cufin    = sum(_ssf(x) for x in s6 if (x['ag'] or '')[:3] == '612')
    gas_total = sum(_ssf(x) for x in s6 if (x['ag'] or '')[:3] not in ('612', '613', '614'))
    return dict(nd_recargos=recargos, nd_multas=multas, nd_permanente=recargos + multas,
                nd_sinreq=sinreq, nd_cufin=cufin, gas_total=gas_total,
                isr_prov=saldo_g.get('114', 0.0),
                hay_multas_recargos=(abs(recargos) >= 1 or abs(multas) >= 1))

FISCAL_CAMPOS = ["util_fiscal_estimada", "coef_utilidad", "iva_devolucion_tramite"]

def get_captura_fiscal(cli, periodo):
    """Lee la captura fiscal manual de un cliente/periodo. Devuelve dict clave->valor."""
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT clave, valor FROM captura_fiscal WHERE cliente_id=%s AND periodo=%s", (cli, periodo))
        return {c: (float(v) if v is not None else None) for c, v in cur.fetchall()}
    except Exception:
        return {}
    finally:
        cur.close(); conn.close()

def save_captura_fiscal(cli, periodo, data):
    """Upsert de la captura fiscal manual. data: dict clave->valor (None se omite)."""
    conn = get_conn(); cur = conn.cursor()
    try:
        for k, v in data.items():
            if v is None or v == "":
                continue
            cur.execute("""INSERT INTO captura_fiscal (cliente_id,periodo,clave,valor,actualizado)
                           VALUES (%s,%s,%s,%s,now())
                           ON CONFLICT (cliente_id,periodo,clave)
                           DO UPDATE SET valor=EXCLUDED.valor, actualizado=now()""",
                        (cli, periodo, k, float(v)))
        conn.commit()
    finally:
        cur.close(); conn.close()

def _meses_iva_favor(cli, periodo, n=3):
    """Cuenta meses consecutivos (hacia atrás desde periodo) con IVA neto a favor. Requiere serie. R-FIS-06."""
    import datetime as _dt
    y, m = int(str(periodo)[:4]), int(str(periodo)[5:7])
    cont = 0
    conn = get_conn(); cur = conn.cursor()
    try:
        for _ in range(n):
            per = _dt.date(y, m, 1)
            try:
                cur.execute("SELECT valor FROM fn_eiva(%s,%s) WHERE concepto=%s",
                            (cli, per, 'IVA neto (+cargo / -favor)'))
                r = cur.fetchone()
            except Exception:
                r = None
            if r and r[0] is not None and float(r[0]) < 0:
                cont += 1
            else:
                break
            m -= 1
            if m == 0:
                m = 12; y -= 1
    finally:
        cur.close(); conn.close()
    return cont

def _sem_c2(fis):
    """Control 2 — no deducible. Automático desde contabilidad. R-FIS-08."""
    gt = fis["gas_total"] or 0
    perm = fis["nd_permanente"]; pp = (perm / gt * 100) if gt else 0
    if fis["hay_multas_recargos"]:
        return ("rojo", "C2 · No deducible", money(perm),
                "Multas/recargos contabilizados: no deducible permanente y falla de cumplimiento.",
                "Diagnóstico de causa raíz + Programa Cero Recargos")
    if pp > 3:
        return ("amarillo", "C2 · No deducible", "{:.1f}%".format(pp),
                "No deducible permanente > 3% del gasto.", "Blindaje de Deducciones")
    if (fis["nd_sinreq"] or 0) > 0:
        return ("amarillo", "C2 · No deducible", money(fis["nd_sinreq"]),
                "Hay no deducible sin requisitos identificado.", "Blindaje de Deducciones")
    return ("verde", "C2 · No deducible", "{:.1f}%".format(pp), "Sin multas/recargos en el periodo.", None)

def _sem_c3(isr_prov, util_fiscal_est):
    """Control 3 — tasa efectiva ISR. Compuerta de pérdida. R-FIS-05."""
    if util_fiscal_est is None:
        return ("gris", "C3 · Tasa efectiva ISR", "s/d", "Falta capturar la utilidad fiscal estimada.", "Captura fiscal")
    if util_fiscal_est <= 0:
        return ("neutral", "C3 · Tasa efectiva ISR", "no aplica",
                "Utilidad fiscal proyectada <= 0: no hay ISR que optimizar.", None)
    isr_proy = util_fiscal_est * 0.30
    bp = ((isr_prov or 0) - isr_proy) / isr_proy * 100 if isr_proy else 0
    if abs(bp) <= 10:
        return ("verde", "C3 · Tasa efectiva ISR", "{:+.0f}%".format(bp),
                "Pagos provisionales alineados al ISR proyectado.", None)
    if abs(bp) <= 25:
        lbl = "sobre-pago" if bp > 0 else "sub-provisión"
        return ("amarillo", "C3 · Tasa efectiva ISR", "{:+.0f}%".format(bp),
                "Brecha de " + lbl + " moderada.", "Revisar pagos provisionales")
    if bp > 25:
        return ("rojo", "C3 · Tasa efectiva ISR", "{:+.0f}%".format(bp),
                "Sobre-pago material: caja atrapada en el SAT.", "Optimización de pagos provisionales (reducción)")
    return ("rojo", "C3 · Tasa efectiva ISR", "{:+.0f}%".format(bp),
            "Sub-provisión material: shock fiscal + recargos 2.07% en puerta.", "Ajuste de provisión inmediato")

def _sem_c4(iva_neto, meses_favor, flag_tramite):
    """Control 4 — IVA a favor / caja atrapada. R-FIS-09."""
    a_favor = (iva_neto is not None and iva_neto < 0)
    mag = money(abs(iva_neto)) if iva_neto is not None else "s/d"
    if not a_favor or flag_tramite:
        return ("verde", "C4 · IVA a favor", mag, "Sin saldo a favor, o devolución en trámite.", None)
    if meses_favor >= 3:
        return ("rojo", "C4 · IVA a favor", mag,
                "Saldo a favor " + str(meses_favor) + " meses sin trámite: caja atrapada.",
                "Recuperación de IVA (devolución, LIVA art. 6)")
    return ("amarillo", "C4 · IVA a favor", mag, "Saldo a favor reciente sin acción.", "Recuperación de IVA")

def evaluar_fiscal(cli, periodo):
    """Evalúa la sección fiscal (C2 auto, C3/C4 con captura). Devuelve lista de semáforos. C1 en stand-by."""
    res = comparativo_estados(cli, periodo)
    if not res:
        return None
    ef = res[0]; fis = ef["fiscal"]
    cap = get_captura_fiscal(cli, periodo)
    # C4: IVA neto del periodo
    iva_neto = None
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT valor FROM fn_eiva(%s,%s) WHERE concepto=%s",
                    (cli, periodo, 'IVA neto (+cargo / -favor)'))
        r = cur.fetchone()
        if r and r[0] is not None:
            iva_neto = float(r[0])
    except Exception:
        iva_neto = None
    finally:
        cur.close(); conn.close()
    meses_favor = _meses_iva_favor(cli, periodo) if (iva_neto is not None and iva_neto < 0) else 0
    flag = bool(cap.get("iva_devolucion_tramite"))
    return [
        _sem_c2(fis),
        _sem_c3(fis["isr_prov"], cap.get("util_fiscal_estimada")),
        _sem_c4(iva_neto, meses_favor, flag),
    ]

def estados_financieros(archivo_id):
    """ER (NIF) + Balance General + EFE (indirecto) por partida. Lógica validada en datos reales."""
    from collections import defaultdict
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT i.cod_agrupador, i.naturaleza::text, i.saldo_final,
                   r.saldo_inicial, r.debe, r.haber, i.num_cuenta
            FROM insumos_balanza i
            JOIN raw_balanza r ON r.archivo_id=i.archivo_id AND r.num_cuenta=i.num_cuenta
            WHERE i.archivo_id=%s AND i.es_hoja AND NOT i.es_orden
        """, (archivo_id,))
        rows = cur.fetchall()
        cur.execute("SELECT cliente_id, periodo FROM insumos_balanza WHERE archivo_id=%s LIMIT 1", (archivo_id,))
        meta = cur.fetchone()
        ene_rows = None
        if meta:
            import datetime as _dt
            cur.execute("SELECT archivo_vigente FROM periodo_estado WHERE cliente_id=%s AND periodo=%s",
                        (meta[0], _dt.date(meta[1].year, 1, 1)))
            _en = cur.fetchone()
            if _en and _en[0]:
                cur.execute("""SELECT i.cod_agrupador, i.naturaleza::text, r.saldo_inicial
                               FROM insumos_balanza i
                               JOIN raw_balanza r ON r.archivo_id=i.archivo_id AND r.num_cuenta=i.num_cuenta
                               WHERE i.archivo_id=%s AND i.es_hoja AND NOT i.es_orden""", (_en[0],))
                ene_rows = cur.fetchall()
    finally:
        cur.close(); conn.close()
    acc = [dict(ag=(a or ''), g=(a or '')[:3], nat=(n or ''), sf=float(sf or 0),
                si=float(si or 0), d=float(d or 0), h=float(h or 0), num=nc)
           for a,n,sf,si,d,h,nc in rows]
    def sg(v, nat): return v if nat == 'D' else -v
    # Posiciones firmadas por AGRUPADOR (no por número de cuenta): robusto a renumeración.
    cierre_sig = defaultdict(float)
    for _x in acc: cierre_sig[_x['ag']] += sg(_x['sf'], _x['nat'])
    apertura_sig = defaultdict(float)
    if ene_rows is not None:
        base_anual = True
        for _a, _n, _si in ene_rows: apertura_sig[_a or ''] += sg(float(_si or 0), _n or '')
    else:
        base_anual = False
        for _x in acc: apertura_sig[_x['ag']] += sg(_x['si'], _x['nat'])
    def _yd(ag): return cierre_sig.get(ag, 0.0) - apertura_sig.get(ag, 0.0)
    p1 = lambda x: x['ag'][:1]
    DEP = ('159','171','172','613','614')
    # ---- Estado de Resultados (todo 4/5/6/7) ----
    ing = sum(x['h']-x['d'] for x in acc if p1(x)=='4')
    cos = sum(x['d']-x['h'] for x in acc if p1(x)=='5')
    seis = [x for x in acc if p1(x)=='6']
    dep = sum(x['d']-x['h'] for x in seis if x['g'] in DEP)
    gas = sum(x['d']-x['h'] for x in seis) - dep
    fin = sum(x['d']-x['h'] for x in acc if p1(x)=='7')
    uai = ing-cos-gas-dep-fin
    er = dict(ing=ing, cos=cos, ub=ing-cos, gas=gas, dep=dep, fin=fin, uai=uai)
    # ---- firmas ----
    ssf = lambda x: x['sf'] if x['nat']=='D' else -x['sf']
    res_ytd = -sum(ssf(x) for x in acc if p1(x) in ('4','5','6','7'))
    ing_y = -sum(ssf(x) for x in acc if p1(x)=='4')
    cos_y = sum(ssf(x) for x in acc if p1(x)=='5')
    seis_y = [x for x in acc if p1(x)=='6']
    dep_y = sum(ssf(x) for x in seis_y if x['g'] in DEP)
    gas_y = sum(ssf(x) for x in seis_y) - dep_y
    fin_y = sum(ssf(x) for x in acc if p1(x)=='7')
    ebit_y = ing_y - cos_y - gas_y - dep_y
    ytd = dict(ing=ing_y, cos=cos_y, ub=ing_y-cos_y, gas=gas_y, dep=dep_y, fin=fin_y, ebit=ebit_y, uai=ebit_y-fin_y)
    # ---- Balance General por partida (3 dígitos) ----
    saldo_g = defaultdict(float)
    for x in acc:
        if p1(x) in ('1','2','3'): saldo_g[x['g']] += ssf(x)
    def bg_lineas(cond, signo):
        return [(_lbl(g), signo*v) for g,v in sorted(saldo_g.items()) if cond(g) and abs(v)>=1]
    act_circ   = bg_lineas(lambda g: g[:1]=='1' and g < '150', 1)
    act_nocirc = bg_lineas(lambda g: g[:1]=='1' and g >= '150', 1)
    pas_cp     = bg_lineas(lambda g: g[:1]=='2' and g < '250', -1)
    pas_lp     = bg_lineas(lambda g: g[:1]=='2' and g >= '250', -1)
    cap        = bg_lineas(lambda g: g[:1]=='3', -1)
    cap.append(('Resultado del ejercicio', res_ytd))
    tot_act = sum(v for g,v in saldo_g.items() if g[:1]=='1')
    tot_pas = sum(-v for g,v in saldo_g.items() if g[:1]=='2')
    tot_cap = sum(-v for g,v in saldo_g.items() if g[:1]=='3') + res_ytd
    bg = dict(act_circ=act_circ, act_nocirc=act_nocirc, pas_cp=pas_cp, pas_lp=pas_lp, cap=cap,
              tot_act=tot_act, tot_pas=tot_pas, tot_cap=tot_cap, residual=tot_act-(tot_pas+tot_cap),
              saldo_g=dict(saldo_g), res=res_ytd)
    # ---- EFE por partida (clasifica por agrupador completo, presenta por 3 dígitos) ----
    def actividad(ag):
        if ag[:3] in ('101','102','103'): return 'EF'
        if ag.startswith(('202.01','205.06','107.05','251','252','253','254','255','256','257','258','259',
                          '301','302','303','304','305')): return 'FIN'
        if ag[:3] in ('150','151','152','153','154','155','156','157','158','159','171','172','184'): return 'INV'
        if ag[:1] in ('1','2'): return 'OP'
        if ag[:1]=='3': return 'FIN'
        return 'OTRO'
    efe_acc = defaultdict(float)
    for ag in set(cierre_sig) | set(apertura_sig):
        if ag[:1] in ('4','5','6','7'): continue      # resultados -> entran vía UAI
        k = actividad(ag)
        if k == 'EF': continue                          # efectivo -> se reporta como saldo, no como flujo
        efe_acc[(k, ag[:3])] += -_yd(ag)                # efecto en caja = -(cambio firmado del agrupador)
    show = lambda K: [(_lbl(g), v) for (k,g),v in sorted(efe_acc.items()) if k==K and abs(v)>=1]
    op_l, inv_l, fin_l = show('OP'), show('INV'), show('FIN')
    op_sum  = sum(v for (k,g),v in efe_acc.items() if k=='OP')
    inv_tot = sum(v for (k,g),v in efe_acc.items() if k=='INV')
    fin_tot = sum(v for (k,g),v in efe_acc.items() if k=='FIN')
    op_tot  = ytd['uai'] + op_sum
    efec_ini = sum(v for ag,v in apertura_sig.items() if ag[:3] in ('101','102','103'))
    efec_fin = sum(v for ag,v in cierre_sig.items()   if ag[:3] in ('101','102','103'))
    dcash = efec_fin - efec_ini
    suma = op_tot + inv_tot + fin_tot
    plug = efec_fin - (efec_ini + suma)
    # R-EFE-01: el flujo cuadra a tolerancia de redondeo o se retiene.
    ap_descuadre = sum(apertura_sig.values())                                  # apertura del ejercicio debe sumar ~0
    ap_resultado = sum(v for ag,v in apertura_sig.items() if ag[:1] in ('4','5','6','7'))  # !=0 => ejercicio anterior sin cerrar
    if base_anual:
        # R-EFE-01: cuadra solo si el flujo cierra, la apertura suma cero y los resultados de apertura están en cero.
        cuadra = abs(plug) < 1.0 and abs(ap_descuadre) < 1.0 and abs(ap_resultado) < 1.0
    else:
        cuadra = abs(plug) < 1.0   # sin enero cargado: solo cuadre mensual; la apertura del ejercicio no se evalúa
    efe = dict(uai=ytd['uai'], op_l=op_l, inv_l=inv_l, fin_l=fin_l, op_sum=op_sum,
               op_tot=op_tot, inv_tot=inv_tot, fin_tot=fin_tot,
               suma=suma, dcash=dcash, efec_ini=efec_ini, efec_fin=efec_fin,
               plug=plug, cuadra=cuadra, ap_descuadre=ap_descuadre, ap_resultado=ap_resultado,
               base_anual=base_anual,
               acc={("%s|%s" % (k, g)): v for (k, g), v in efe_acc.items()})
    # ---- ISR provisional: cuenta contable, agrupador 114 "Pagos provisionales" (leído de balanza) ----
    AG_ISR_PROV = '114'
    isr_prov = sum(x['sf'] for x in acc if x['g'] == AG_ISR_PROV)
    ytd['isr_prov'] = isr_prov
    ytd['un'] = ytd['uai'] - isr_prov
    # ---- Generador de Efectivo (Cash Effectiveness, Alexander) · YTD · por agrupador ----
    capex = sum(_yd(ag) for ag in cierre_sig if '150' <= ag[:3] <= '184' and ag[:3] not in ('159', '171', '172'))
    def _es_opcap_ag(ag):
        cur_act = ag[:1] == '1' and ag[:3] < '150'
        cur_pas = ag[:1] == '2' and ag[:3] < '250'
        return (cur_act or cur_pas) and actividad(ag) == 'OP'
    opcap     = sum(v for ag,v in cierre_sig.items()   if _es_opcap_ag(ag))
    opcap_ini = sum(v for ag,v in apertura_sig.items() if _es_opcap_ag(ag))
    dopcap = opcap - opcap_ini
    generador = ytd['uai'] + ytd['dep'] - capex - dopcap
    # Inventario inmóvil del ejercicio: agrupador de inventario (115-129) cuya posición no se movió desde la apertura.
    inv_inmovil = sum(v for ag,v in cierre_sig.items() if '115' <= ag[:3] <= '129' and abs(_yd(ag)) < 1)
    cash = dict(uai=ytd['uai'], dep=ytd['dep'], capex=capex, opcap=opcap, opcap_ini=opcap_ini,
                dopcap=dopcap, generador=generador, isr_prov=isr_prov, un=ytd['un'], inv_inmovil=inv_inmovil)
    return dict(er=er, bg=bg, efe=efe, ytd=ytd, cash=cash, fiscal=_fiscal_buckets(acc, saldo_g))

def _fmt(v):
    if v is None: return "—"
    return ("(${:,.0f})".format(abs(v))) if v < 0 else ("${:,.0f}".format(v))

def _cmp_md(rows, col_a, col_p):
    out = ["| Concepto | " + col_a + " | " + col_p + " | Variación |", "|---|--:|--:|--:|"]
    for label, va, vp, bold in rows:
        if va is None and vp is None:
            out.append("| **" + label + "** | | | |")
        else:
            var = (va or 0) - (vp or 0)
            if bold:
                out.append("| **" + label + "** | **" + _fmt(va) + "** | " + _fmt(vp) + " | **" + _fmt(var) + "** |")
            else:
                out.append("| " + label + " | " + _fmt(va) + " | " + _fmt(vp) + " | " + _fmt(var) + " |")
    return "\n".join(out)

def comparativo_estados(cli, per_actual):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT archivo_vigente FROM periodo_estado WHERE cliente_id=%s AND periodo=%s", (cli, per_actual))
        r = cur.fetchone(); bal_a = r[0] if r else None
        cur.execute("""SELECT periodo, archivo_vigente FROM periodo_estado
                       WHERE cliente_id=%s AND periodo<%s AND archivo_vigente IS NOT NULL
                       ORDER BY periodo DESC LIMIT 1""", (cli, per_actual))
        rp = cur.fetchone()
    finally:
        cur.close(); conn.close()
    if not bal_a: return None
    ef_a = estados_financieros(bal_a)
    ef_p = estados_financieros(rp[1]) if rp else None
    per_p = str(rp[0])[:7] if rp else None
    return ef_a, ef_p, per_p

def _bg_sec(g):
    if g[:1] == '1': return ('Activo circulante', 1) if g < '150' else ('Activo no circulante', 1)
    if g[:1] == '2': return ('Pasivo corto plazo', -1) if g < '250' else ('Pasivo largo plazo', -1)
    return ('Capital contable', -1)

def er_rows_cmp(ef_a, ef_p):
    A = ef_a['ytd']; p = ef_p['ytd'] if ef_p else {}
    L = [('Ingresos','ing',False),('(−) Costo de ventas','cos',False),('= Utilidad bruta','ub',True),
         ('(−) Gastos de operación','gas',False),('(−) Depreciación','dep',False),
         ('= Utilidad de operación','ebit',True),
         ('(−) Resultado financiero neto','fin',False),('= Utilidad antes de impuestos','uai',True)]
    return [(lbl, A.get(k), p.get(k), b) for lbl, k, b in L]

def bg_rows_cmp(ef_a, ef_p):
    sa = ef_a['bg']['saldo_g']; sp = ef_p['bg']['saldo_g'] if ef_p else {}
    codes = sorted(set(sa) | set(sp))
    rows = []
    for sec in ['Activo circulante','Activo no circulante']:
        cs = [g for g in codes if _bg_sec(g)[0] == sec]
        if cs:
            rows.append((sec, None, None, False))
            for g in cs:
                va = sa.get(g,0)*_bg_sec(g)[1]; vp = sp.get(g,0)*_bg_sec(g)[1]
                if abs(va) >= 1 or abs(vp) >= 1: rows.append((_lbl(g), va, vp, False))
    rows.append(('= Total activo', ef_a['bg']['tot_act'], (ef_p['bg']['tot_act'] if ef_p else None), True))
    for sec in ['Pasivo corto plazo','Pasivo largo plazo']:
        cs = [g for g in codes if _bg_sec(g)[0] == sec]
        if cs:
            rows.append((sec, None, None, False))
            for g in cs:
                va = sa.get(g,0)*-1; vp = sp.get(g,0)*-1
                if abs(va) >= 1 or abs(vp) >= 1: rows.append((_lbl(g), va, vp, False))
    rows.append(('= Total pasivo', ef_a['bg']['tot_pas'], (ef_p['bg']['tot_pas'] if ef_p else None), True))
    rows.append(('Capital contable', None, None, False))
    for g in [g for g in codes if g[:1] == '3']:
        va = sa.get(g,0)*-1; vp = sp.get(g,0)*-1
        if abs(va) >= 1 or abs(vp) >= 1: rows.append((_lbl(g), va, vp, False))
    rows.append(('Resultado del ejercicio', ef_a['bg']['res'], (ef_p['bg']['res'] if ef_p else None), False))
    rows.append(('= Total capital', ef_a['bg']['tot_cap'], (ef_p['bg']['tot_cap'] if ef_p else None), True))
    rows.append(('= Pasivo + Capital (debe igualar el activo)',
                 ef_a['bg']['tot_pas'] + ef_a['bg']['tot_cap'],
                 ((ef_p['bg']['tot_pas'] + ef_p['bg']['tot_cap']) if ef_p else None), True))
    return rows

def efe_rows_cmp(ef_a, ef_p):
    aa = ef_a['efe']['acc']; ap = ef_p['efe']['acc'] if ef_p else {}
    keys = sorted(set(aa) | set(ap))
    P = ef_p['efe'] if ef_p else {}
    rows = []
    for sec, code in [('FLUJO DE OPERACIÓN','OP'),('FLUJO DE INVERSIÓN','INV'),('FLUJO DE FINANCIAMIENTO','FIN')]:
        rows.append((sec, None, None, False))
        if code == 'OP':
            rows.append(('Utilidad antes de impuestos', ef_a['efe']['uai'], P.get('uai'), False))
        for k in keys:
            if k.split('|')[0] != code: continue
            va = aa.get(k,0); vp = ap.get(k,0)
            if abs(va) >= 1 or abs(vp) >= 1:
                rows.append(('Δ ' + _lbl(k.split('|')[1]), va, vp, False))
        tk = {'OP':'op_tot','INV':'inv_tot','FIN':'fin_tot'}[code]
        rows.append(('= ' + sec.replace('FLUJO DE ','Flujo de ').lower().capitalize(), ef_a['efe'][tk], P.get(tk), True))
    rows.append(('= Variación neta de efectivo', ef_a['efe']['suma'], P.get('suma'), True))
    rows.append(('(+) Efectivo al inicio del ejercicio', ef_a['efe']['efec_ini'], P.get('efec_ini'), False))
    rows.append(('= Efectivo al final (calculado)', ef_a['efe']['efec_ini']+ef_a['efe']['suma'],
                 ((P.get('efec_ini') or 0)+(P.get('suma') or 0)) if ef_p else None, True))
    rows.append(('Efectivo al final (real, Caja + Bancos)', ef_a['efe']['efec_fin'], P.get('efec_fin'), False))
    return rows


def _fr(v, fmt):
    if v is None: return "—"
    if fmt == 'pct':   return "{:.1f}%".format(v*100)
    if fmt == 'dias':  return "{:,.0f} días".format(v)
    if fmt == 'veces': return "{:.2f}x".format(v)
    if fmt == 'pesos': return _fmt(v)
    return "{:.2f}".format(v)

def _fvar(va, vp, fmt):
    if va is None or vp is None: return "—"
    d = va - vp
    if fmt == 'pct':   return "{:+.1f} pp".format(d*100)
    if fmt == 'dias':  return "{:+,.0f} días".format(d)
    if fmt == 'veces': return "{:+.2f}x".format(d)
    if fmt == 'pesos': return _fmt(d)
    return "{:+.2f}".format(d)

def _stocks_de(ef):
    sg = ef['bg']['saldo_g']
    act_circ = sum(v for c,v in sg.items() if c[:1]=='1' and c < '150')
    pas_circ = sum(-v for c,v in sg.items() if c[:1]=='2' and c < '250')
    cxc = sg.get('105',0) + sg.get('106',0)
    inv = sg.get('115',0)+sg.get('116',0)+sg.get('117',0)+sg.get('121',0)
    cxp = -sg.get('201',0)
    afn = sum(v for c,v in sg.items() if '150' <= c < '180')
    return dict(act_circ=act_circ, pas_circ=pas_circ, cxc=cxc, inv=inv, cxp=cxp, afn=afn,
                activo=ef['bg']['tot_act'], pasivo=ef['bg']['tot_pas'], capital=ef['bg']['tot_cap'])

# ---------------------------------------------------------------------------
# CAPA META (R-MET) · metas fijas por ejercicio + elegibilidad de licitacion
# ---------------------------------------------------------------------------
def get_metas(cli, ejercicio):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""SELECT indicador, tipo, direccion, valor_meta, fuente, nota
                       FROM meta_indicador WHERE cliente_id=%s AND ejercicio=%s""", (cli, ejercicio))
        return {r[0]: dict(tipo=r[1], direccion=r[2],
                           valor_meta=(float(r[3]) if r[3] is not None else None),
                           fuente=r[4], nota=r[5]) for r in cur.fetchall()}
    finally:
        cur.close(); conn.close()

def set_meta(cli, ejercicio, indicador, valor_meta, tipo='umbral', direccion='mayor_mejor', fuente=None, nota=None):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""INSERT INTO meta_indicador (cliente_id,ejercicio,indicador,tipo,direccion,valor_meta,fuente,nota)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (cliente_id,ejercicio,indicador)
                       DO UPDATE SET tipo=EXCLUDED.tipo, direccion=EXCLUDED.direccion, valor_meta=EXCLUDED.valor_meta,
                                     fuente=EXCLUDED.fuente, nota=EXCLUDED.nota, actualizado=now()""",
                    (cli, ejercicio, indicador, tipo, direccion, valor_meta, fuente, nota))
        conn.commit()
    finally:
        cur.close(); conn.close()

# Constantes federales de licitacion (R-MET-03): se cablean, no se guardan en BD.
LICIT_AC_PC = 1.1     # AC/PC minimo
LICIT_AT_PT = 2.0     # AT/PT minimo
LICIT_PT_AT = 0.70    # PT/AT maximo
LICIT_CNT_PCT = 0.20  # CNT minimo = 20% de la propuesta sin IVA

def licitacion_eval(ef, propuesta):
    """Razones del balance vs parametros federales. Elegibilidad = a AND (b OR c). R-MET-03."""
    s = _stocks_de(ef)
    AC, PC, AT, PT = s['act_circ'], s['pas_circ'], s['activo'], s['pasivo']
    cnt   = AC - PC
    ac_pc = (AC / PC) if PC else None
    at_pt = (AT / PT) if PT else None
    pt_at = (PT / AT) if AT else None
    cnt_meta = (LICIT_CNT_PCT * propuesta) if propuesta else None
    a = (cnt >= cnt_meta) if cnt_meta is not None else None
    b = (ac_pc is not None and ac_pc >= LICIT_AC_PC) and (at_pt is not None and at_pt >= LICIT_AT_PT)
    c = (pt_at is not None and pt_at <= LICIT_PT_AT)
    elegible = (bool(a) and (bool(b) or bool(c))) if a is not None else None
    return dict(AC=AC, PC=PC, AT=AT, PT=PT, cnt=cnt, cnt_meta=cnt_meta,
                ac_pc=ac_pc, at_pt=at_pt, pt_at=pt_at, a=a, b=b, c=c,
                elegible=elegible, propuesta=propuesta)

# Anclas duras (Crabtree) cableadas como default; meta_indicador las sobrescribe por cliente. R-SU-16.
DEFAULT_METAS = {
    'pretax_pct': dict(tipo='umbral', direccion='mayor_mejor', valor_meta=10.0, fuente='crabtree',
                       nota='Pre-Tax >=10% (Crabtree, piso).'),
    'gpld':       dict(tipo='umbral', direccion='mayor_mejor', valor_meta=1.35, fuente='crabtree',
                       nota='GPLD >=1.35 (Crabtree).'),
    'cash_lag':   dict(tipo='umbral', direccion='menor_mejor', valor_meta=0.0,  fuente='crabtree',
                       nota='Cash Lag <=0: la utilidad llega a caja.'),
}

def meta_de(indicador, metas):
    """Meta sembrada por cliente (meta_indicador) tiene prioridad sobre el default cableado."""
    return metas.get(indicador) or DEFAULT_METAS.get(indicador)

def evaluar_meta(valor, m):
    """Devuelve (brecha_firmada, estado). Brecha negativa = falta para llegar a la meta. R-MET-02."""
    if valor is None or not m or m.get('valor_meta') is None:
        return None, None
    meta = m['valor_meta']; dirn = m.get('direccion', 'mayor_mejor')
    signo = -1 if dirn == 'menor_mejor' else 1
    brecha = (valor - meta) * signo
    if dirn == 'contextual':
        return brecha, 'contextual'
    return brecha, ('cumple' if brecha >= 0 else 'falta')

def _fmeta(v, fmt):
    if v is None: return "—"
    if fmt == 'pctnum': return "{:.1f}%".format(v)
    if fmt == 'x':      return "{:.2f}x".format(v)
    if fmt == 'pesos':  return _fmt(v)
    return "{:.2f}".format(v)

def _fmeta_brecha(v, fmt):
    if v is None: return "—"
    if fmt == 'pctnum': return "{:+.1f} pp".format(v)
    if fmt == 'x':      return "{:+.2f}x".format(v)
    if fmt == 'pesos':  return _fmt(v)
    return "{:+.2f}".format(v)

def tabla_avance_meta(items, metas):
    """items = [(label, indicador_key, valor, fmt)]. Tabla markdown Actual/Meta/Brecha/Estado, o None si nada tiene meta."""
    rows = ["| Indicador | Actual | Meta | Brecha | Estado |", "|---|--:|--:|--:|:--|"]
    edo_txt = {'cumple': 'Cumple', 'falta': 'Falta', 'contextual': 'Contextual'}
    hay = False
    for label, key, valor, fmt in items:
        m = meta_de(key, metas)
        if not m or m.get('valor_meta') is None:
            continue
        hay = True
        brecha, estado = evaluar_meta(valor, m)
        rows.append("| " + label + " | " + _fmeta(valor, fmt) + " | " + _fmeta(m['valor_meta'], fmt) +
                    " | " + _fmeta_brecha(brecha, fmt) + " | " + edo_txt.get(estado, "—") + " |")
    return "\n".join(rows) if hay else None

def ratios_mensuales(cli, periodo):
    import datetime
    sd = lambda a,b: (a/b) if b else None
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT periodo, archivo_vigente FROM periodo_estado WHERE cliente_id=%s AND archivo_vigente IS NOT NULL ORDER BY periodo", (cli,))
        rowsp = cur.fetchall()
    finally:
        cur.close(); conn.close()
    pers = [str(p) for p,_ in rowsp]; arch = {str(p): a for p,a in rowsp}
    if periodo not in pers: return None
    cache = {}
    def ef_de(idx):
        if idx < 0 or idx >= len(pers): return None
        per = pers[idx]
        if per not in cache: cache[per] = estados_financieros(arch[per])
        return cache[per]
    def dias_ytd(per):
        y,m = int(per[:4]), int(per[5:7])
        d = datetime.date(y,12,31) if m==12 else datetime.date(y,m+1,1)-datetime.timedelta(days=1)
        return d.timetuple().tm_yday
    def snap(idx):
        ef = ef_de(idx)
        if ef is None: return None
        ytd = ef['ytd']; st = _stocks_de(ef); per = pers[idx]
        ing=ytd['ing']; cos=ytd['cos']; gas=ytd['gas']; dep=ytd['dep']; fin=ytd['fin']; ebit=ytd['ebit']; uai=ytd['uai']; un=ytd['un']; cash=ef['cash']
        dias = dias_ytd(per)
        efp = ef_de(idx-1)
        crec = (ef['er']['ing']/efp['er']['ing'] - 1) if (efp and efp['er']['ing']) else None
        dso = sd(st['cxc'], ing); dso = dso*dias if dso is not None else None
        dio = sd(st['inv'], cos); dio = dio*dias if dio is not None else None
        dpo = sd(st['cxp'], cos); dpo = dpo*dias if dpo is not None else None
        cce = (dso+dio-dpo) if None not in (dso,dio,dpo) else None
        wc = st['act_circ'] - st['pas_circ']
        return [
            ('1. Evaluación Operativa','Crecimiento en ingresos (MoM)', crec, 'pct'),
            ('1. Evaluación Operativa','Margen de utilidad bruta', sd(ing-cos,ing), 'pct'),
            ('1. Evaluación Operativa','Margen de gastos de operación', sd(gas,ing), 'pct'),
            ('1. Evaluación Operativa','Margen de utilidad operativa (EBIT)', sd(ebit,ing), 'pct'),
            ('1. Evaluación Operativa','Margen de utilidad antes de impuestos', sd(uai,ing), 'pct'),
            ('1. Evaluación Operativa','Margen de utilidad neta', sd(un,ing), 'pct'),
            ('1. Evaluación Operativa','Utilidad antes de impuestos (UAI)', uai, 'pesos'),
            ('1. Evaluación Operativa','(−) ISR provisional (cuenta 114)', cash['isr_prov'], 'pesos'),
            ('1. Evaluación Operativa','Utilidad neta', un, 'pesos'),
            ('2. Uso de Activos','Días cuentas por cobrar (DSO)', dso, 'dias'),
            ('2. Uso de Activos','Rotación de inventario', sd(cos, st['inv']), 'veces'),
            ('2. Uso de Activos','Días de inventario (DIO)', dio, 'dias'),
            ('2. Uso de Activos','Días cuentas por pagar (DPO)', dpo, 'dias'),
            ('2. Uso de Activos','Ciclo de conversión de efectivo', cce, 'dias'),
            ('2. Uso de Activos','Capital de trabajo', wc, 'pesos'),
            ('2. Uso de Activos','Capital de trabajo / Ingresos', sd(wc, ing), 'pct'),
            ('2. Uso de Activos','Rotación de activos fijos', sd(ing, st['afn']), 'veces'),
            ('2. Uso de Activos','Rotación de activo total', sd(ing, st['activo']), 'veces'),
            ('3. Estructura de Capital y Liquidez','Liquidez (razón corriente)', sd(st['act_circ'], st['pas_circ']), 'veces'),
            ('3. Estructura de Capital y Liquidez','Liquidez inmediata (prueba ácida)', sd(st['act_circ']-st['inv'], st['pas_circ']), 'veces'),
            ('3. Estructura de Capital y Liquidez','Deuda / Capitalización', sd(st['pasivo'], st['pasivo']+st['capital']), 'pct'),
            ('3. Estructura de Capital y Liquidez','Deuda / Capital contable', sd(st['pasivo'], st['capital']), 'veces'),
            ('3. Estructura de Capital y Liquidez','Cobertura de intereses', sd(ebit, fin), 'veces'),
            ('4. Rentabilidad','Retorno sobre activos (ROA)', sd(un, st['activo']), 'pct'),
            ('4. Rentabilidad','Retorno sobre capital (ROE)', sd(un, st['capital']), 'pct'),
            ('4. Rentabilidad','DuPont — Margen neto', sd(un,ing), 'pct'),
            ('4. Rentabilidad','DuPont — Rotación de activo', sd(ing, st['activo']), 'veces'),
            ('4. Rentabilidad','DuPont — Apalancamiento (Activo/Capital)', sd(st['activo'], st['capital']), 'veces'),
            ('4. Rentabilidad','DuPont — ROE (producto)',
                ((sd(un,ing) or 0)*(sd(ing,st['activo']) or 0)*(sd(st['activo'],st['capital']) or 0)) if st['capital'] else None, 'pct'),
            ('4. Rentabilidad','ROIC (ROCE de libro)', sd(ebit, st['activo']-st['pas_circ']), 'pct'),
            ('5. Generador de Efectivo','Utilidad antes de impuestos', cash['uai'], 'pesos'),
            ('5. Generador de Efectivo','(+) Depreciación y amortización', cash['dep'], 'pesos'),
            ('5. Generador de Efectivo','(−) Gastos de capital (CapEx)', cash['capex'], 'pesos'),
            ('5. Generador de Efectivo','(−) Incremento de capital operativo', cash['dopcap'], 'pesos'),
            ('5. Generador de Efectivo','= Generador de efectivo (YTD)', cash['generador'], 'pesos'),
            ('5. Generador de Efectivo','Generador de efectivo / Ingresos', sd(cash['generador'], ing), 'pct'),
            ('5. Generador de Efectivo','Capital operativo (posición)', cash['opcap'], 'pesos'),
            ('5. Generador de Efectivo','(−) Inventario inmóvil del ejercicio', cash['inv_inmovil'], 'pesos'),
            ('5. Generador de Efectivo','= Capital operativo vivo (sin inv. inmóvil)', cash['opcap'] - cash['inv_inmovil'], 'pesos'),
            ('5. Generador de Efectivo','Capital de trabajo (posición, referencia)', wc, 'pesos'),
            ('5. Generador de Efectivo','= Capital de trabajo vivo (sin inv. inmóvil)', wc - cash['inv_inmovil'], 'pesos'),
        ]
    iM = pers.index(periodo)
    return snap(iM), snap(iM-1), (pers[iM-1] if iM > 0 else None)

# ---------------------------------------------------------------------------
# INTERFAZ
# ---------------------------------------------------------------------------
st.title("📊 Tablero CFO Vastion")
st.caption("Carga mensual · arrastra los archivos del cliente y valida")

try:
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id, nombre FROM cliente ORDER BY nombre")
    clientes = cur.fetchall(); cur.close(); conn.close()
except Exception as e:
    st.error(f"No se pudo conectar a la base. Revisa los secrets. ({e})"); st.stop()

if not clientes:
    st.warning("No hay clientes registrados."); st.stop()

nombre_sel = st.selectbox("Cliente", [c[1] for c in clientes])
cli_id = dict((c[1], c[0]) for c in clientes)[nombre_sel]

with st.expander("Datos del cliente · Constancia de Situación Fiscal", expanded=False):
    _saved_csf = get_cliente_csf(cli_id) or {}
    if st.session_state.get("csf_cli") != cli_id:
        for _k in CSF_CAMPOS:
            st.session_state["csf_" + _k] = _saved_csf.get(_k, "") or ""
        st.session_state["csf_cli"] = cli_id
        st.session_state["csf_last"] = None
    _csf_pdf = st.file_uploader("Sube la constancia (PDF) para autocompletar los campos", type=["pdf"], key="csf_up")
    if _csf_pdf is not None and st.session_state.get("csf_last") != _csf_pdf.name:
        try:
            _ext = parse_constancia(_csf_pdf.getvalue())
            for _k in CSF_CAMPOS:
                if _ext.get(_k): st.session_state["csf_" + _k] = _ext[_k]
            st.session_state["csf_last"] = _csf_pdf.name
            st.success("Constancia leída. Revisa y corrige los campos antes de guardar.")
        except Exception as e:
            st.warning(f"No pude leer la constancia automáticamente ({e}). Captura los campos a mano.")
    _ca, _cb = st.columns(2)
    with _ca:
        st.text_input("Razón social", key="csf_razon_social")
        st.text_input("RFC", key="csf_rfc")
        st.text_input("Nombre comercial", key="csf_nombre_comercial")
        st.text_input("Régimen capital", key="csf_regimen_capital")
        st.text_input("Actividad económica", key="csf_actividad_economica")
        st.text_input("Porcentaje actividad", key="csf_actividad_pct")
    with _cb:
        st.text_input("Régimen fiscal", key="csf_regimen_fiscal")
        st.text_input("Estatus en el padrón", key="csf_estatus")
        st.text_input("Inicio de operaciones", key="csf_fecha_inicio_ops")
        st.text_input("Código postal", key="csf_cp")
        st.text_input("Municipio", key="csf_municipio")
        st.text_input("Entidad federativa", key="csf_entidad")
    if st.button("Guardar datos del cliente"):
        try:
            save_cliente_csf(cli_id, {_k: st.session_state.get("csf_" + _k, "") for _k in CSF_CAMPOS})
            st.success("Datos del cliente guardados.")
        except Exception as e:
            st.error(f"No se pudieron guardar (¿corriste cliente_csf.sql en Supabase?). {e}")

st.markdown("**Suelta aquí los archivos del mes** (balanza XML obligatoria; catálogo XML la primera vez; "
            "CFDI emitidos y recibidos; auxiliar de Clientes para cartera/DSO):")
subidos = st.file_uploader("Arrastra los archivos", accept_multiple_files=True,
                           type=['xml','xlsx'], label_visibility='collapsed')

archivos = {}
if subidos:
    for f in subidos:
        data = f.getvalue(); tipo = clasificar(f.name, data)
        archivos[tipo] = (f.name, data)
    etiquetas = {'BALANZA':'Balanza','CATALOGO':'Catálogo','CFDI_EMITIDO':'CFDI emitidos','CFDI_RECIBIDO':'CFDI recibidos','AUX_CLIENTES':'Auxiliar Clientes'}
    st.info("Detecté: " + ", ".join(etiquetas.get(t, t) for t in archivos))
    if 'BALANZA' not in archivos:
        if 'AUX_CLIENTES' in archivos:
            st.info("Para cargar solo auxiliares de Clientes (sin balanza), usa el cargador de la sección "
                    "**Cartera y DSO por cliente** más abajo. En esta carga mensual la balanza es obligatoria.")
        else:
            st.error("Falta la balanza (XML). Es obligatoria.")

if st.button("Cargar y validar", type="primary", disabled=not (subidos and 'BALANZA' in archivos)):
    with st.spinner("Cargando y validando…"):
        try:
            per, bal_id, integ, madz = procesar(cli_id, archivos)
            ind = cargar_indicadores(bal_id, cli_id, per)
        except Exception as e:
            st.error(f"Error al procesar: {e}"); st.stop()

    bloqueo = any((not ok) and sev=='BLOQUEANTE' for _,ok,sev,_ in integ) or any((not ok) and sev=='BLOQUEANTE' for _,ok,sev,_ in madz)
    adv = any((not ok) and sev=='ADVERTENCIA' for _,ok,sev,_ in integ) or any((not ok) and sev=='ADVERTENCIA' for _,ok,sev,_ in madz)

    st.divider()
    if bloqueo:
        st.error(f"## 🔴 RETENIDO · {per}\nEste periodo no entra al tablero. Corregir en contabilidad antes de reportar.")
    elif adv:
        st.warning(f"## 🟡 VALIDADO con advertencias · {per}\nEl periodo entra, pero revisa las señales amarillas.")
    else:
        st.success(f"## 🟢 VALIDADO · {per}\nEl periodo entra al tablero.")

    st.caption("Indicadores clave, apoyo y avance contra meta: ve la sección **Indicadores clave** abajo (elige el mes).")

    # ---- detalle de compuertas ----
    with st.expander("Detalle de validación"):
        st.markdown("**Integridad**")
        for p, ok, sev, det in integ:
            st.write((("✅" if ok else ("🟡" if sev=='ADVERTENCIA' else "❌"))) + f"  **{p}** — {det}")
        st.markdown("**Madurez analítica**")
        for p, ok, sev, det in madz:
            st.write((("✅" if ok else ("🟡" if sev=='ADVERTENCIA' else "❌"))) + f"  **{p}** ({sev}) — {det}")


# ---------------------------------------------------------------------------
# 1 - CAPTURA MANUAL (onboarding + mensual). Lo que el sistema no puede calcular.
# ---------------------------------------------------------------------------
st.divider()
st.markdown("### 1 · Captura manual")
st.caption("Llena esto primero. Solo lo que no vive en la contabilidad; el resto lo lee el sistema del agrupador. "
           "Los datos del cliente (Constancia) se capturan arriba, al elegir el cliente.")

with st.expander("Metas del ejercicio (onboarding) — propuesta de licitación", expanded=False):
    _pmeta = periodos_cargados(cli_id)
    _ej0 = int(_pmeta[0][:4]) if _pmeta else 2026
    _ejm = st.number_input("Ejercicio", value=_ej0, step=1, format="%d", key="meta_ej")
    _metas_cur = get_metas(cli_id, int(_ejm))
    _prow = _metas_cur.get("licit_propuesta")
    _prv = _prow["valor_meta"] if _prow else 0.0
    st.caption("Un solo dato fija la meta de licitación del ejercicio: el CNT meta = 20% de la propuesta (sin IVA).")
    _nv = st.number_input("Propuesta objetivo del ejercicio (sin IVA, MXN)",
                          min_value=0.0, value=float(_prv), step=1_000_000.0, format="%.0f", key="meta_prop")
    _isn_row = _metas_cur.get("isn_tasa")
    _isn_v = _isn_row["valor_meta"] if _isn_row else 4.0
    _isn = st.number_input("Tasa de ISN del estado (%) — NL 2026 = 4.0; se parametriza por estado (R-FIS-07)",
                           min_value=0.0, max_value=10.0, value=float(_isn_v), step=0.5, format="%.2f", key="meta_isn")
    if st.button("Guardar meta del ejercicio", key="meta_save"):
        set_meta(cli_id, int(_ejm), "licit_propuesta", _nv, tipo="umbral", direccion="mayor_mejor",
                 fuente="licitacion", nota="Propuesta objetivo sin IVA; CNT meta = 20% de este valor.")
        set_meta(cli_id, int(_ejm), "isn_tasa", _isn, tipo="parametro", direccion="contextual",
                 fuente="ley_hacienda_estatal", nota="Tasa ISN del estado del cliente; se siembra en onboarding.")
        st.success("Meta y tasa ISN guardadas para el ejercicio " + str(int(_ejm)) + ".")

with st.expander("Identidad visual del cliente (logo y colores)", expanded=False):
    _brand = get_branding(cli_id) or {}
    st.caption("Viste el reporte del cliente. El semáforo (rojo/amarillo/verde) no se tiñe: es universal y no se toca.")
    _bc1, _bc2 = st.columns(2)
    with _bc1:
        _cprim = st.color_picker("Color primario (títulos)", value=_brand.get("color_primario") or "#1C2D3A", key="brand_prim")
    with _bc2:
        _cacc = st.color_picker("Color de acento (línea/detalles)", value=_brand.get("color_acento") or "#2E86AB", key="brand_acc")
    _logo_up = st.file_uploader("Logo del cliente (PNG o JPG, preferente fondo claro)", type=["png", "jpg", "jpeg"], key="brand_logo")
    if _brand.get("logo"):
        st.caption("Logo actual: " + (_brand.get("logo_nombre") or "guardado") + ". Sube uno nuevo solo si quieres reemplazarlo.")
    if st.button("Guardar identidad visual", key="brand_save"):
        try:
            _lb = _logo_up.getvalue() if _logo_up is not None else None
            _ln = _logo_up.name if _logo_up is not None else None
            save_branding(cli_id, _cprim, _cacc, _lb, _ln)
            st.success("Identidad visual guardada.")
        except Exception as e:
            st.error("No se pudo guardar (¿corriste cliente_branding.sql en Supabase?). " + repr(e))

with st.expander("Captura fiscal del mes (Controles 3 y 4)", expanded=False):
    _pcf = periodos_cargados(cli_id)
    if not _pcf:
        st.caption("Carga al menos un mes para capturar la parte fiscal mensual.")
    else:
        _mcf = st.selectbox("Mes", _pcf, key="capfis_mes")
        _capf = get_captura_fiscal(cli_id, _mcf)
        st.caption("Utilidad fiscal estimada y coeficiente (C3); bandera de devolución de IVA (C4).")
        _cca, _ccb = st.columns(2)
        with _cca:
            _ufe = st.number_input("Utilidad fiscal estimada del ejercicio (acumulada, C3)",
                                   value=float(_capf.get("util_fiscal_estimada") or 0.0), step=10000.0, key="capfis_ufe")
            _coef = st.number_input("Coeficiente de utilidad (informativo, C3)",
                                    value=float(_capf.get("coef_utilidad") or 0.0), step=0.01, format="%.4f", key="capfis_coef")
        with _ccb:
            _tram = st.checkbox("Devolución/compensación de IVA en trámite (C4)",
                                value=bool(_capf.get("iva_devolucion_tramite")), key="capfis_tram")
        if st.button("Guardar captura fiscal", key="capfis_save"):
            try:
                save_captura_fiscal(cli_id, _mcf, {
                    "util_fiscal_estimada": _ufe if _ufe else None,
                    "coef_utilidad": _coef if _coef else None,
                    "iva_devolucion_tramite": 1 if _tram else 0})
                st.success("Captura fiscal guardada.")
            except Exception as e:
                st.error("No se pudo guardar (¿corriste captura_fiscal.sql en Supabase?). " + repr(e))


# ---------------------------------------------------------------------------
# INDICADORES CLAVE (revision por mes)
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Indicadores clave")
st.caption("Revisión por mes. Caja al centro (regla de hierro). Tres ejes: generación, traslado, fiscal.")
_pind = periodos_cargados(cli_id)
if not _pind:
    st.caption("Carga al menos un mes.")
else:
    mes_ind = st.selectbox("Mes", _pind, key="ind_mes")
    if st.button("Ver indicadores", key="ind_btn"):
        _cn = get_conn(); _cu = _cn.cursor()
        _cu.execute("SELECT archivo_vigente FROM periodo_estado WHERE cliente_id=%s AND periodo=%s", (cli_id, mes_ind))
        _rb = _cu.fetchone(); _cu.close(); _cn.close()
        if not _rb or not _rb[0]:
            st.error("No se encontró la balanza de ese mes.")
        else:
            ind = cargar_indicadores(_rb[0], cli_id, mes_ind)
            c1, c2, c3 = st.columns(3)
            with c1:
                st.caption("EJE 1 · GENERACIÓN")
                st.metric("Pre-Tax Profit %", f"{ind['pretax_pct']}%" if ind['pretax_pct'] is not None else "—")
                st.caption(sem_pretax(ind['pretax_pct']))
            with c2:
                st.caption("EJE 2 · TRASLADO")
                st.metric("Cash Lag", money(ind['cash_lag']))
                if ind['cash_lag'] is None:
                    st.caption("Necesita mes previo")
                elif ind['cash_lag'] > 0:
                    st.caption("🔴 Caja cayó más que la utilidad")
                else:
                    st.caption("🟢 Caja por encima de la utilidad")
            with c3:
                st.caption("EJE 3 · FISCAL")
                st.metric("EIVA %", f"{ind['eiva_pct']}%" if ind['eiva_pct'] is not None else "—")
                if ind['iva_neto'] is not None:
                    st.caption(("IVA a cargo " + money(ind['iva_neto'])) if ind['iva_neto'] > 0
                               else ("IVA a favor " + money(abs(ind['iva_neto']))))
            st.subheader("Apoyo")
            a1, a2, a3, a4 = st.columns(4)
            a1.metric("Margen Bruto %", f"{ind['mb_pct']}%" if ind['mb_pct'] is not None else "—")
            a2.metric("Nómina / Ingresos", f"{ind['nomina_pct']}%" if ind['nomina_pct'] is not None else "—")
            a3.metric("GPLD", f"{ind['gpld']}x" if ind['gpld'] is not None else "n/a")
            a4.metric("Caja fin de mes", money(ind['caja']))
            st.subheader("Avance contra meta")
            _metas_b = get_metas(cli_id, int(mes_ind[:4]))
            _items_b = [("Pre-Tax Profit %", "pretax_pct", ind["pretax_pct"], "pctnum"),
                        ("GPLD",             "gpld",       ind["gpld"],       "x"),
                        ("Cash Lag",         "cash_lag",   ind["cash_lag"],   "pesos")]
            _tab_b = tabla_avance_meta(_items_b, _metas_b)
            if _tab_b:
                st.markdown(_tab_b)
                st.caption("Brecha firmada: negativa = falta para llegar a la meta (R-MET-02). "
                           "Anclas duras Pre-Tax 10% y GPLD 1.35 (Crabtree); se sobrescriben por cliente desde meta_indicador.")
            else:
                st.caption("Sin metas definidas para este cliente/ejercicio.")
            with st.expander("P&L de Gestión (no constituye estado de resultados NIF)"):
                st.dataframe(
                    [{"Concepto": c, "Monto": f"{m:,.0f}", "%": (f"{p}%" if p is not None else "")} for c, m, p in ind['pl']],
                    use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# TOP 10 CLIENTES (CFDI EMITIDOS) - alimenta Alerta #3 (concentracion de cliente)
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Top 10 Clientes · CFDI emitidos")
st.caption("Ingreso neto facturado (sin IVA, vigente, neto de notas de credito). "
           "Llave = RFC. Alimenta la Alerta #3 (concentracion de cliente).")
_ptc = periodos_cargados(cli_id)
if not _ptc:
    st.caption("Carga al menos un mes de CFDI emitidos.")
else:
    mes_tc = st.selectbox("Mes", _ptc, key="tc_mes")
    if st.button("Ver Top 10", key="tc_btn"):
        data = top_clientes(cli_id, mes_tc)
        for clave, titulo in (("mes", "Del mes"), ("acumulado", "Acumulado del ejercicio (YTD)")):
            d = data[clave]
            st.markdown(f"**{titulo}** — {d['clientes']} clientes · ingreso neto {money(d['total'])}")
            if d["conc_top1"] is not None:
                luz = "🔴" if d["conc_top1"] >= 50 else ("🟡" if d["conc_top1"] >= 30 else "🟢")
                st.caption(f"{luz} Concentracion · cliente #1 = {d['conc_top1']}% · Top 3 = {d['conc_top3']}%")
            if d["top"]:
                st.dataframe(
                    [{"#": i, "Cliente": t["cliente"], "RFC": t["rfc"],
                      "Ingreso neto": money(t["neto"]),
                      "%": (f"{t['pct']}%" if t["pct"] is not None else ""),
                      "Fac": t["facturas"]} for i, t in enumerate(d["top"], 1)],
                    use_container_width=True, hide_index=True)
            else:
                st.caption("Sin CFDI emitidos en el rango.")


# ---------------------------------------------------------------------------
# MEZCLA DE INGRESOS / MODELO DE NEGOCIO (CFDI emitidos por segmento SAT)
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Mezcla de ingresos · Modelo de negocio")
st.caption("Composicion del ingreso emitido por segmento SAT (2 primeros digitos de la Clave SAT), leida de los "
           "CFDI ya cargados. Ingreso neto = Subtotal - Descuento (sin IVA), solo Vigente, neto de notas de credito. "
           "YTD del ejercicio del mes seleccionado. No requiere subir nada: usa lo que cargaste por el cargador normal.")
_pmz = periodos_cfdi_emitido(cli_id)
if not _pmz:
    st.caption("Aun no hay CFDI emitidos cargados para este cliente.")
else:
    mes_mz = st.selectbox("Mes (la mezcla es YTD del ejercicio de ese mes)", _pmz, key="mz_mes")
    if st.button("Ver mezcla", key="mz_btn", type="primary"):
        mz = mezcla_ingresos(cli_id, mes_mz)
        st.caption(f"Comprobantes emitidos vigentes en el ejercicio {mz['ejercicio']}: {mz['total_cfdi']} · "
                   f"con Clave SAT: {mz['con_clave']} ({mz['cob_clave']}%).")
        if mz["total_cfdi"] == 0:
            st.warning("No hay CFDI emitidos vigentes cargados en ese ejercicio.")
        elif mz["con_clave"] == 0:
            st.warning("Los CFDI de ese ejercicio se cargaron antes de capturar la Clave SAT y, por la inmutabilidad "
                       "de raw_cfdi, no se reescriben. Los que cargues de aqui en adelante ya entran con su clave y "
                       "apareceran solos.")
        elif not mz["rows"]:
            st.warning("No hay ingreso emitido vigente en ese ejercicio.")
        else:
            d = mz["dominante"]
            st.markdown(f"**Modelo de negocio: {d['etiqueta']} ({d['pct']}%)** · "
                        f"ingreso YTD {money(mz['total'])} · {len(mz['rows'])} segmento(s).")
            st.caption("El segmento dominante define la capa especifica. Dos segmentos que cubren el grueso = "
                       "dos lineas; uno solo >80% = monolinea.")
            st.dataframe(
                [{"Segmento": "[" + r["seg"] + "] " + r["etiqueta"],
                  "Ingreso YTD": money(r["neto"]),
                  "%": (f"{r['pct']}%" if r["pct"] is not None else ""),
                  "Fac": r["facturas"]} for r in mz["rows"]],
                use_container_width=True, hide_index=True)
            if mz["pct_sin_clave"] > 0:
                st.caption(f"Nota: {mz['pct_sin_clave']}% del ingreso del ejercicio aun sin Clave SAT "
                           "(comprobantes cargados antes de la captura).")


st.caption("Saldo por cobrar de cada cliente, desde la BALANZA (fuente autoritativa; reconcilia por construccion). "
           "Ordenado por exposicion: por aqui empieza la cobranza. El ultimo movimiento (del auxiliar) marca la "
           "cartera congelada: saldo grande + sin movimiento reciente = dinero estancado, prioridad uno.")
with st.expander("Cargar auxiliar(es) de Clientes — solo para fechas de movimiento · opcional · no requiere balanza", expanded=False):
    st.caption("El saldo de cartera NO depende de este archivo (sale de la balanza). El auxiliar solo aporta la fecha "
               "del ultimo movimiento por cliente. Sube el auxiliar de pólizas de la 105; puedes subir varios a la vez.")
    _auxup = st.file_uploader("Auxiliar(es) de Clientes (.xlsx)", type=['xlsx'],
                              accept_multiple_files=True, key="aux_up")
    if _auxup and st.button("Cargar auxiliares", key="aux_btn"):
        try:
            _res = cargar_auxiliares(cli_id, [(f.name, f.getvalue()) for f in _auxup])
            for nom, ej, n in _res:
                if ej:
                    st.success(f"{nom}: {n} movimientos cargados.")
                else:
                    st.error(f"{nom}: no se reconocio como auxiliar de Clientes (105). Revisa el archivo.")
        except Exception as e:
            st.error(f"Error al cargar (¿corriste la migracion raw_aux_clientes en Supabase?): {e}")
_pdso = periodos_cargados(cli_id)
if not _pdso:
    st.caption("Carga al menos un mes.")
else:
    mes_dso = st.selectbox("Mes", _pdso, key="dso_mes")
    if st.button("Ver cartera", key="dso_btn"):
        dd = cartera_clientes(cli_id, mes_dso)
        if not dd["rows"]:
            st.warning("La balanza de ese mes no tiene saldo en cuentas de Clientes (105). Revisa que la balanza este cargada.")
        else:
            if dd["estim_ok"]:
                st.markdown(f"**Cartera bruta {money(dd['total'])} − Estimacion {money(dd['estim'])} "
                            f"= Cartera neta {money(dd['neto'])}** · cobertura de incobrables {dd['cobertura']}% · "
                            f"{len(dd['rows'])} clientes.")
                st.caption("Bruta = saldo de Clientes (105) en balanza, reconcilia por construccion. "
                           "Neta = bruta menos la estimacion para cuentas incobrables (agrupador 108). "
                           "El capital de trabajo se calcula sobre la bruta (NIF C-3); la neta es lo que se vuelve caja.")
            else:
                st.markdown(f"**Cartera bruta: {money(dd['total'])}** · {len(dd['rows'])} clientes con saldo · "
                            f"fuente: balanza (cuenta Clientes 105).")
                st.warning("Sin estimacion para cuentas incobrables agrupada (108.01). Cartera neta no disponible. "
                           "La estimacion existe en libros pero no tiene codigo agrupador, asi que es invisible por "
                           "agrupador (y es una exposicion ante el SAT en el envio de balanza, CFF 28). "
                           "Corregir en Contabilidad: asignar agrupador 108.01 a la cuenta de estimacion en Contalink, "
                           "reexpedir catalogo y balanza. El neteo aparece solo cuando este agrupada.")
            st.caption("Saldo de balanza (reconcilia por construccion). DSO = saldo / facturado 12 meses moviles x 365. "
                       "CONGELADA = saldo sin facturacion en 12 meses. Dias antiguedad = dias desde el ultimo "
                       "movimiento al cierre del mes: nunca queda vacio y mide que tan vieja es la cartera. "
                       "s/dato = sin auxiliar cargado para esa cuenta.")
            def _dso_txt(v):
                if v is None: return "s/dato"
                if v == "CONGELADA": return "CONGELADA"
                return f"{v:.0f}"
            st.dataframe(
                [{"Cliente": o["cliente"],
                  "Cartera": money(o["saldo"]),
                  "%": (f"{o['pct']}%" if o["pct"] is not None else ""),
                  "DSO": _dso_txt(o["dso"]),
                  "Dias antiguedad": (str(o["antiguedad"]) if o["antiguedad"] is not None else "s/dato"),
                  "Facturado 12m": money(o["fac12"]),
                  "Ultimo movimiento": (o["ult_mov"] or "sin dato")} for o in dd["rows"]],
                use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# VELOCIDAD DE COBRO (auxiliar de Clientes, FIFO factura->cobro) - complementa la cartera
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Velocidad de cobro \u00b7 dias factura -> cobro")
st.caption("Cuanto tarda en cobrarse cada peso que SI se cobro, desde el auxiliar de Clientes (fuente que "
           "reconcilia con la balanza). No usa la Fecha de pago del CFDI: ese campo no es fiable (da dias "
           "negativos y marca como pagados a clientes que la cartera tiene congelados). Los congelados no "
           "salen aqui porque no cobraron; viven en la Cartera de arriba. Ventana: cobros del ejercicio.")
_pvc = periodos_cargados(cli_id)
if not _pvc:
    st.caption("Carga al menos un mes.")
else:
    mes_vc = st.selectbox("Mes (ejercicio del cobro)", _pvc, key="vc_mes")
    if st.button("Ver velocidad de cobro", key="vc_btn"):
        vc = velocidad_cobro(cli_id, mes_vc)
        if not vc["rows"]:
            st.warning("No hay cobros conciliados en el auxiliar para ese ejercicio. "
                       "Sube el auxiliar de Clientes (105) del ejercicio en la seccion de arriba.")
        else:
            st.markdown(f"**Velocidad de cobro agregada: {vc['agregado']} dias** \u00b7 "
                        f"sobre {money(vc['tot_cobrado'])} cobrado y emparejado en {vc['ejercicio']} \u00b7 "
                        f"{len(vc['rows'])} clientes con cobro.")
            st.caption("Emparejamiento FIFO: cada cobro salda las facturas mas viejas primero. Complementa la "
                       "Cartera: alli ves quien te debe (saldo); aqui, que tan rapido cobras a quien si paga.")
            st.dataframe(
                [{"Cliente": o["cliente"], "Cuenta": o["sub"],
                  "Dias de cobro": o["dias"],
                  "Cobrado " + str(vc["ejercicio"]): money(o["cobrado"])} for o in vc["rows"]],
                use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# CONTROL 1 - FRACCIONAMIENTO / RIESGO 69-B (CFDI emitidos) - patron de materialidad
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Control 1 \u00b7 Fraccionamiento (riesgo 69-B)")
st.caption("Patron de materialidad en CFDI emitidos: 2+ comprobantes vigentes el mismo dia al mismo receptor, "
           "cada uno menor a $1M pero sumando $1M o mas. Son pedazos sub-millon que reconstruyen una operacion "
           "que habria cruzado el umbral. SEÑAL DE RIESGO, no veredicto: obra publica se factura por estimacion "
           "y puede explicar el patron. Marca para revisar. Ventana: YTD del ejercicio del mes.")
_pfr = periodos_cfdi_emitido(cli_id)
if not _pfr:
    st.caption("Aun no hay CFDI emitidos cargados para este cliente.")
else:
    mes_fr = st.selectbox("Mes (ejercicio a revisar)", _pfr, key="fr_mes")
    if st.button("Detectar fraccionamiento", key="fr_btn", type="primary"):
        fr = fraccionamiento(cli_id, mes_fr)
        if not fr["rows"]:
            st.success(f"Sin patron de fraccionamiento en {fr['ejercicio']}: no hay dias con 2+ CFDI al mismo "
                       "receptor, cada uno bajo $1M y sumando $1M o mas.")
        else:
            st.markdown(f"**{fr['n_grupos']} grupos marcados** \u00b7 {fr['n_cfdi']} CFDI \u00b7 "
                        f"suma {money(fr['suma_total'])} \u00b7 ejercicio {fr['ejercicio']}.")
            st.caption("Cada renglon es un dia+receptor a revisar. Prioriza receptores privados y recurrentes: "
                       "un mismo cliente que reaparece varios dias del año pesa mas que un municipio con una "
                       "estimacion de obra puntual. Cruza contra la Cartera: un receptor aqui que ademas este "
                       "CONGELADO alli apila señales.")
            st.dataframe(
                [{"Fecha": o["dia"], "Receptor": o["cliente"], "RFC": o["rfc"],
                  "# CFDI": o["n"], "Suma del dia": money(o["suma"])} for o in fr["rows"]],
                use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# PROVEEDORES 69-B (CFDI RECIBIDOS) - fraccionamiento + edad del proveedor por RFC
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Proveedores 69-B \u00b7 CFDI recibidos")
st.caption("Riesgo de materialidad en proveedores (Art. 69-B CFF): deducir o acreditar un CFDI de un EFOS te "
           "tumba la deduccion y el IVA. Dos criterios: (A) fraccionamiento -mismo dia, mismo proveedor, cada "
           "CFDI < $1M y sumando $1M o mas-; (B) edad del proveedor sacada del RFC -recien constituido facturando "
           "fuerte es bandera clasica-. SEÑAL DE RIESGO, no veredicto. Ventana: YTD del ejercicio del mes.")
_pr = periodos_cfdi_recibido(cli_id)
if not _pr:
    st.caption("Aun no hay CFDI recibidos cargados para este cliente.")
else:
    mes_pr = st.selectbox("Mes (ejercicio a revisar)", _pr, key="pr69_mes")
    if st.button("Evaluar proveedores 69-B", key="pr69_btn", type="primary"):
        pr = proveedores_69b(cli_id, mes_pr)
        st.markdown(f"**Criterio A \u00b7 Fraccionamiento** \u2014 ejercicio {pr['ejercicio']}")
        if not pr["frac_rows"]:
            st.success("Sin patron de fraccionamiento en proveedores este ejercicio.")
        else:
            st.caption("Proveedor que te factura partido en pedazos sub-millon el mismo dia. "
                       "La columna Edad marca los recien constituidos: fraccionamiento + proveedor joven apila señales.")
            st.dataframe(
                [{"Fecha": o["dia"], "Proveedor": o["proveedor"], "RFC": o["rfc"],
                  "Tipo": o["tipo"], "Edad (años)": ("s/dato" if o["edad"] is None else o["edad"]),
                  "# CFDI": o["n"], "Suma del dia": money(o["suma"])} for o in pr["frac_rows"]],
                use_container_width=True, hide_index=True)
        st.markdown(f"**Criterio B \u00b7 Proveedores jovenes** (menos de {pr['edad_max']:.0f} años, con gasto) "
                    f"\u2014 {len(pr['jovenes'])} de {pr['n_prov']} proveedores")
        if not pr["jovenes"]:
            st.success(f"Ningun proveedor con menos de {pr['edad_max']:.0f} años y gasto relevante este ejercicio.")
        else:
            st.caption("Proveedor recien constituido no es delito, pero un RFC joven concentrando gasto es de lo "
                       "primero que revisa el SAT. Cruza contra la lista 69-B publicada cuando la carguemos.")
            st.dataframe(
                [{"Proveedor": o["proveedor"], "RFC": o["rfc"], "Tipo": o["tipo"],
                  "Constituido": o["constitucion"], "Edad (años)": o["edad"],
                  "Gasto YTD": money(o["gasto"]), "# CFDI": o["ncfdi"]} for o in pr["jovenes"]],
                use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# ESTADOS FINANCIEROS FORMALES (NIF) - REPORTE MENSUAL
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Estados financieros formales (NIF) — Reporte Mensual")
st.caption("ER y EFE en acumulado del ejercicio (YTD); Balance en posición del mes. Para clientes con observaciones, son diagnóstico hasta la recontabilización.")
_pf = periodos_cargados(cli_id)
if not _pf:
    st.caption("Carga al menos un mes para generar los estados financieros.")
else:
    mes_ef = st.selectbox("Mes", _pf, key="ef_mes")
    comparar = st.checkbox("Comparar con el mes inmediato anterior", key="ef_comp")
    if st.button("Generar estados financieros"):
        res = comparativo_estados(cli_id, mes_ef)
        if res is None:
            st.error("No se encontró la balanza de ese mes.")
        else:
            ef_a, ef_p, per_p = res
            if comparar and ef_p is None:
                st.warning("No hay un mes anterior cargado para comparar. Mostrando solo el mes seleccionado.")
                comparar = False
            ca = mes_ef[:7]
            ok_efe, causa_efe, acc_efe = _efe_diag(ef_a["efe"])
            if not ok_efe:
                st.error("ESTADOS RETENIDOS (R-EFE-01) - " + causa_efe + "  **Accion:** " + acc_efe +
                         "  Los estados de abajo son diagnostico interno, no entregables al cliente.")
            if comparar:
                cp = per_p
                st.markdown("#### Estado de Resultados · " + ca + " vs " + cp + " (acumulado)")
                st.markdown(_cmp_md(er_rows_cmp(ef_a, ef_p), ca, cp))
                st.markdown("#### Balance General · " + ca + " vs " + cp)
                st.markdown(_cmp_md(bg_rows_cmp(ef_a, ef_p), ca, cp))
                rsd = ef_a["bg"]["residual"]
                if abs(rsd) < 1: st.success("Balance del mes cuadrado (residual " + _fmt(rsd) + ").")
                else: st.warning("Residual de cuadre del mes: " + _fmt(rsd) + ".")
                st.markdown("#### Estado de Flujo de Efectivo · " + ca + " vs " + cp + " (indirecto, acumulado)")
                st.markdown(_cmp_md(efe_rows_cmp(ef_a, ef_p), ca, cp))
                plg = ef_a["efe"]["plug"]
                if abs(plg) < 1: st.success("EFE del mes cuadrado (partida por identificar $0).")
                else: st.warning("Partida por identificar del mes: " + _fmt(plg) + ".")
                st.caption("La columna Variación es el cambio de cada cuenta entre los dos meses — el mismo insumo que alimenta el flujo de efectivo.")
            else:
                er = ef_a["ytd"]; bg = ef_a["bg"]; efe = ef_a["efe"]
                pct = lambda v: (" ({:.1f}%)".format(v/er["ing"]*100)) if er["ing"] else ""
                st.markdown("#### Estado de Resultados · " + ca + " (acumulado del ejercicio)")
                _er = ["| Concepto | Monto |", "|---|--:|",
                       "| Ingresos | " + _fmt(er["ing"]) + " |",
                       "| (−) Costo de ventas | " + _fmt(er["cos"]) + " |",
                       "| **= Utilidad bruta** | **" + _fmt(er["ub"]) + "**" + pct(er["ub"]) + " |",
                       "| (−) Gastos de operación | " + _fmt(er["gas"]) + " |",
                       "| (−) Depreciación | " + _fmt(er["dep"]) + " |",
                       "| **= Utilidad de operación (EBIT)** | **" + _fmt(er["ebit"]) + "**" + pct(er["ebit"]) + " |",
                       "| (−) Resultado financiero neto | " + _fmt(er["fin"]) + " |",
                       "| **= Utilidad antes de impuestos** | **" + _fmt(er["uai"]) + "**" + pct(er["uai"]) + " |"]
                st.markdown(chr(10).join(_er))
                filas = ["| Partida | Monto |", "|---|--:|", "| **ACTIVO CIRCULANTE** | |"]
                for l,v in bg["act_circ"]: filas.append("| " + l + " | " + _fmt(v) + " |")
                filas.append("| **ACTIVO NO CIRCULANTE** | |")
                for l,v in bg["act_nocirc"]: filas.append("| " + l + " | " + _fmt(v) + " |")
                filas.append("| **= Total activo** | **" + _fmt(bg["tot_act"]) + "** |")
                filas.append("| **PASIVO CORTO PLAZO** | |")
                for l,v in bg["pas_cp"]: filas.append("| " + l + " | " + _fmt(v) + " |")
                if bg["pas_lp"]:
                    filas.append("| **PASIVO LARGO PLAZO** | |")
                    for l,v in bg["pas_lp"]: filas.append("| " + l + " | " + _fmt(v) + " |")
                filas.append("| **= Total pasivo** | **" + _fmt(bg["tot_pas"]) + "** |")
                filas.append("| **CAPITAL CONTABLE** | |")
                for l,v in bg["cap"]: filas.append("| " + l + " | " + _fmt(v) + " |")
                filas.append("| **= Total capital** | **" + _fmt(bg["tot_cap"]) + "** |")
                filas.append("| Pasivo + Capital | " + _fmt(bg["tot_pas"]+bg["tot_cap"]) + " |")
                st.markdown("#### Balance General · " + ca)
                st.markdown("\n".join(filas))
                if abs(bg["residual"]) < 1: st.success("Balance cuadrado (residual " + _fmt(bg["residual"]) + ").")
                else: st.warning("Residual de cuadre: " + _fmt(bg["residual"]) + " — saldos desfasados (R-EC-10 / naturaleza).")
                fil = ["| Concepto | Monto |", "|---|--:|", "| **OPERACIÓN** | |",
                       "| Utilidad antes de impuestos | " + _fmt(efe["uai"]) + " |"]
                for l,v in efe["op_l"]: fil.append("| Δ " + l + " | " + _fmt(v) + " |")
                fil.append("| **= Flujo de operación** | **" + _fmt(efe["op_tot"]) + "** |")
                fil.append("| **INVERSIÓN** | |")
                for l,v in efe["inv_l"]: fil.append("| Δ " + l + " | " + _fmt(v) + " |")
                fil.append("| **= Flujo de inversión** | **" + _fmt(efe["inv_tot"]) + "** |")
                fil.append("| **FINANCIAMIENTO** | |")
                for l,v in efe["fin_l"]: fil.append("| Δ " + l + " | " + _fmt(v) + " |")
                fil.append("| **= Flujo de financiamiento** | **" + _fmt(efe["fin_tot"]) + "** |")
                fil.append("| **= Variación neta de efectivo** | **" + _fmt(efe["suma"]) + "** |")
                fil.append("| (+) Efectivo y equivalentes al inicio del ejercicio | " + _fmt(efe["efec_ini"]) + " |")
                fil.append("| **= Efectivo y equivalentes al final (calculado)** | **" + _fmt(efe["efec_ini"]+efe["suma"]) + "** |")
                fil.append("| Efectivo y equivalentes al final (real, Caja + Bancos) | " + _fmt(efe["efec_fin"]) + " |")
                fil.append("| Partida por identificar | " + _fmt(efe["plug"]) + " |")
                st.markdown("#### Estado de Flujo de Efectivo · " + ca + " (indirecto, acumulado del ejercicio)")
                st.markdown("\n".join(fil))
                if abs(efe["plug"]) < 1: st.success("EFE cuadrado (partida por identificar $0).")
                else: st.warning("Partida por identificar: " + _fmt(efe["plug"]) + ".")


# ---------------------------------------------------------------------------
# SECCION FISCAL - SEMAFOROS (C2 auto, C3/C4 con captura). C1 stand-by.
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Sección fiscal — semáforos")
st.caption("Tablero delgado: tres semáforos, no indicadores nuevos. C2 (no deducible) sale solo de contabilidad; "
           "C3 y C4 usan captura manual. C1 (69-B/EFOS) en stand-by. Base YTD.")
_pfis = periodos_cargados(cli_id)
if not _pfis:
    st.caption("Carga al menos un mes.")
else:
    mes_fis = st.selectbox("Mes", _pfis, key="fis_mes")
    if st.button("Evaluar sección fiscal", key="fis_eval"):
        _sem = evaluar_fiscal(cli_id, mes_fis)
        if not _sem:
            st.error("No se encontró la balanza de ese mes.")
        else:
            _ico = {"verde": "🟢", "amarillo": "🟡", "rojo": "🔴", "neutral": "⚪", "gris": "⚪"}
            for luz, titulo, valor, msg, accion in _sem:
                _box = {"rojo": st.error, "amarillo": st.warning, "verde": st.success}.get(luz, st.info)
                _txt = _ico.get(luz, "⚪") + " **" + titulo + "** — " + valor + "  \n" + msg
                if accion:
                    _txt += "  \n**Acción facturable:** " + accion
                _box(_txt)
            # detalle de C2 desde contabilidad
            _eff = comparativo_estados(cli_id, mes_fis)[0]["fiscal"]
            st.caption("Detalle C2 (contabilidad): recargos " + money(_eff["nd_recargos"]) +
                       " · multas/sanciones " + money(_eff["nd_multas"]) +
                       " · sin requisitos " + money(_eff["nd_sinreq"]) +
                       " · CUFIN " + money(_eff["nd_cufin"]) +
                       " · ISR provisional pagado " + money(_eff["isr_prov"]))
            st.caption("Nota: la cuenta de ISN de CKT aún se etiqueta '3%'; la ley NL 2026 es 4%. Recosteo de obra pendiente.")


# ---------------------------------------------------------------------------
# RATIOS FINANCIEROS MENSUALES (Alexander) - REPORTE MENSUAL
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Ratios financieros mensuales — Reporte Mensual")
st.caption("Base: acumulado del año a la fecha (YTD), sin anualizar. Los días reflejan los días transcurridos del año.")
_pr = periodos_cargados(cli_id)
if not _pr:
    st.caption("Carga al menos un mes para calcular ratios.")
else:
    mes_r = st.selectbox("Mes", _pr, key="rat_mes")
    if st.button("Calcular ratios"):
        out = ratios_mensuales(cli_id, mes_r)
        if out is None:
            st.error("No se encontró ese mes.")
        else:
            a_rows, p_rows, per_p = out
            pmap = {(s,l):(v,f) for s,l,v,f in p_rows} if p_rows else {}
            ca = mes_r[:7]; cp = per_p[:7] if per_p else "—"
            sec_actual = None; tabla = []
            for s,l,v,f in a_rows:
                if s != sec_actual:
                    if tabla:
                        st.markdown("\n".join(tabla)); tabla = []
                    st.markdown("##### " + s)
                    tabla = ["| Ratio | " + ca + " | " + cp + " | Variación |", "|---|--:|--:|--:|"]
                    sec_actual = s
                vp = pmap.get((s,l),(None,f))[0]
                tabla.append("| " + l + " | " + _fr(v,f) + " | " + _fr(vp,f) + " | " + _fvar(v,vp,f) + " |")
            if tabla: st.markdown("\n".join(tabla))


# ---------------------------------------------------------------------------
# COMPARATIVO DE MESES CARGADOS
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Tendencia de los meses cargados")
if st.button("Ver tendencia"):
    datos = tendencia(cli_id)
    if not datos:
        st.info("Este cliente no tiene meses cargados todavía.")
    else:
        filas = [{"Mes": d["periodo"],
                  "Caja": money(d["caja"]),
                  "Pre-Tax %": (f"{d['pretax']}%" if d["pretax"] is not None else "—"),
                  "Cash Lag": money(d["cash_lag"]),
                  "Comparable": ("⚠️ con observaciones" if d["flag"] else "✓ limpio")} for d in datos]
        st.dataframe(filas, use_container_width=True, hide_index=True)
        import pandas as pd
        caja_df = pd.DataFrame([{"Mes": d["periodo"], "Caja": d["caja"]} for d in datos if d["caja"] is not None]).set_index("Mes")
        if not caja_df.empty:
            st.caption("Caja al cierre por mes")
            st.line_chart(caja_df)
        if any(d["flag"] for d in datos):
            st.caption("⚠️ Los meses marcados traen observaciones de contabilidad (en recontabilización). La tendencia puede no reflejar la operación real hasta que se ajusten.")


# ---------------------------------------------------------------------------
# ELEGIBILIDAD DE LICITACIÓN / CRÉDITO (capa cliente · constructoras) — R-MET
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Elegibilidad de licitación / crédito — constructoras")
st.caption("Capa cliente. Razones del balance contra los parámetros financieros de obra pública federal. "
           "Posición adelantada del mes — el dato certificable es el cierre anual del SAT (R-MET-04).")
_plic = periodos_cargados(cli_id)
if not _plic:
    st.caption("Carga al menos un mes para evaluar elegibilidad.")
else:
    mes_lic = st.selectbox("Mes (posición adelantada)", _plic, key="lic_mes")
    ejercicio = int(mes_lic[:4])
    metas = get_metas(cli_id, ejercicio)
    prop_row = metas.get("licit_propuesta")
    propuesta = prop_row["valor_meta"] if prop_row else None
    if not propuesta:
        st.info("Define la propuesta objetivo del ejercicio en la sección de captura (Metas del ejercicio) para evaluar elegibilidad.")
    else:
        _resl = comparativo_estados(cli_id, mes_lic)
        if not _resl or _resl[0] is None:
            st.error("No se encontró la balanza de ese mes.")
        else:
            ev = licitacion_eval(_resl[0], propuesta)
            edo = lambda ok: "Cumple" if ok else "No cumple"
            vx = lambda v: "—" if v is None else "{:.2f}x".format(v)
            pc = lambda v: "—" if v is None else "{:.1f}%".format(v * 100)
            cnt_brecha = ev["cnt"] - ev["cnt_meta"]
            acpc_b = (ev["ac_pc"] - LICIT_AC_PC) if ev["ac_pc"] is not None else None
            atpt_b = (ev["at_pt"] - LICIT_AT_PT) if ev["at_pt"] is not None else None
            ptat_b = (LICIT_PT_AT - ev["pt_at"]) if ev["pt_at"] is not None else None   # menor-mejor: signo invertido
            filas = ["| Parámetro | Actual | Meta | Brecha | Estado |", "|---|--:|--:|--:|:--|",
                     "| a) Capital neto de trabajo (AC − PC) | " + _fmt(ev["cnt"]) + " | " + _fmt(ev["cnt_meta"]) +
                       " | " + _fmt(cnt_brecha) + " | " + edo(ev["a"]) + " |",
                     "| b) Liquidez · AC / PC | " + vx(ev["ac_pc"]) + " | 1.10x | " +
                       (vx(acpc_b) if acpc_b is not None else "—") + " | " +
                       edo(ev["ac_pc"] is not None and ev["ac_pc"] >= LICIT_AC_PC) + " |",
                     "| b) Solvencia · AT / PT | " + vx(ev["at_pt"]) + " | 2.00x | " +
                       (vx(atpt_b) if atpt_b is not None else "—") + " | " +
                       edo(ev["at_pt"] is not None and ev["at_pt"] >= LICIT_AT_PT) + " |",
                     "| c) Endeudamiento · PT / AT | " + pc(ev["pt_at"]) + " | 70.0% | " +
                       (("{:+.1f} pp".format(ptat_b * 100)) if ptat_b is not None else "—") + " | " +
                       edo(ev["pt_at"] is not None and ev["pt_at"] <= LICIT_PT_AT) + " |"]
            st.markdown("\n".join(filas))
            if ev["elegible"]:
                st.success("ELEGIBLE — cumple a) y al menos uno de b)/c).  Posición adelantada, no certificable.")
            else:
                faltan = []
                if not ev["a"]: faltan.append("a) capital de trabajo")
                if not ev["b"] and not ev["c"]: faltan.append("b) y c)")
                st.error("NO ELEGIBLE — falla " + " · ".join(faltan) +
                         ".  Regla: a AND (b OR c).  Calificación financiera del subrubro: cero.")
            st.caption("Brecha firmada: negativa = falta para llegar a la meta (R-MET-02). "
                       "Diagnóstico sobre la posición del mes; para clientes en recontabilización (Fase 0) "
                       "las razones del balance no son certificables hasta cerrar (R-MET-06).")


# ---------------------------------------------------------------------------
# REPORTE INTERNO (comportamiento del ejercicio · uso Vastion)
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Reporte interno — comportamiento del ejercicio")
st.caption("Uso interno Vastion. Gráficas de tendencia de los indicadores clave + resumen del año "
           "(apertura / cierre / promedio). Para clientes en Fase 0 es diagnóstico, no línea base válida.")
_pint = periodos_cargados(cli_id)
if not _pint:
    st.caption("Carga al menos un mes para generar el reporte interno.")
else:
    _anios = sorted({int(p[:4]) for p in _pint}, reverse=True)
    anio_int = st.selectbox("Ejercicio", _anios, key="int_anio")
    if st.button("Generar reporte interno", key="int_gen"):
        try:
            import matplotlib  # noqa: F401
            _mpl_ok = True
        except Exception:
            _mpl_ok = False
        if not _mpl_ok:
            st.error("Falta la librería **matplotlib**. En GitHub agrega una línea `matplotlib` a requirements.txt, "
                     "guarda, y en Streamlit Cloud entra a *Manage app* y haz *Reboot*.")
        else:
            serie = serie_anual(cli_id, anio_int)
            if not serie:
                st.warning("No hay meses cargados para el ejercicio " + str(anio_int) + ".")
            else:
                _cie = cierre_ejercicio(cli_id, anio_int)
                _efc = None; _diasc = 365; _crec = None; _efp = None; _rn = None; _rp = None
                if _cie:
                    _efc = estados_financieros(_cie[1]); _diasc = _dias_ytd(_cie[0])
                    _rnall = ratios_mensuales(cli_id, _cie[0] + "-01"); _rn = _rnall[0] if _rnall else None
                    _ciep = cierre_ejercicio(cli_id, anio_int - 1)
                    if _ciep:
                        _efp = estados_financieros(_ciep[1])
                        if _efp["ytd"]["ing"]:
                            _crec = _efc["ytd"]["ing"] / _efp["ytd"]["ing"] - 1
                        _rpall = ratios_mensuales(cli_id, _ciep[0] + "-01"); _rp = _rpall[0] if _rpall else None
                _fsem = evaluar_fiscal(cli_id, _cie[0] + "-01") if _cie else None
                try:
                    _pdfint = pdf_interno_2025(nombre_sel, anio_int, serie,
                                               ef_cierre=_efc, dias_cierre=_diasc, crecimiento=_crec,
                                               ef_cierre_prev=_efp, rat_now=_rn, rat_prev=_rp,
                                               datos_cliente=get_cliente_csf(cli_id), fiscal_sem=_fsem)
                    st.download_button("📄 Descargar reporte interno", data=_pdfint,
                                       file_name="Reporte_Interno_" + nombre_sel.replace(" ", "_") + "_" + str(anio_int) + ".pdf",
                                       mime="application/pdf", key="int_dl")
                    st.success("Reporte interno generado con " + str(len(serie)) + " meses del ejercicio " + str(anio_int) + ".")
                except Exception as _e:
                    st.error("No se pudo generar el reporte interno: " + repr(_e))
# ---------------------------------------------------------------------------
# PODER DEL UNO - SIMULADOR DE PALANCAS (Scaling Up / Miltz)
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Poder del Uno — simulador de palancas")
st.caption("Captura cuánto mueves cada palanca y mira el efecto en utilidad y caja, sobre la temporalidad que elijas. "
           "El escenario calculado se pasa al PDF del cliente del mismo mes.")
_ppu = periodos_cargados(cli_id)
if not _ppu:
    st.caption("Carga al menos un mes para usar el simulador.")
else:
    mes_pu = st.selectbox("Mes base", _ppu, key="pu_mes")
    temp_pu = st.radio("Temporalidad", ["mensual", "ytd", "anualizado"], horizontal=True, key="pu_temp",
                       format_func=lambda x: {"mensual": "Del mes", "ytd": "Acumulado del año", "anualizado": "Anualizado"}[x])
    st.caption("Positivo = mejora. Precio/Volumen suben; Costo/Gastos bajan; Días por cobrar/inventario bajan; Días por pagar suben.")
    _cu1, _cu2 = st.columns(2)
    with _cu1:
        pu_precio = st.number_input("Precio (% aumento)", value=1.0, step=0.5, key="pu_precio")
        pu_vol    = st.number_input("Volumen (% aumento)", value=0.0, step=0.5, key="pu_vol")
        pu_costo  = st.number_input("Costo de ventas (% reducción)", value=1.0, step=0.5, key="pu_costo")
        pu_gastos = st.number_input("Gastos de operación (% reducción)", value=1.0, step=0.5, key="pu_gastos")
    with _cu2:
        pu_cxc = st.number_input("Días por cobrar (reducción)", value=0.0, step=5.0, key="pu_cxc")
        pu_inv = st.number_input("Días de inventario (reducción)", value=0.0, step=5.0, key="pu_inv")
        pu_cxp = st.number_input("Días por pagar (aumento)", value=0.0, step=5.0, key="pu_cxp")
    if st.button("Calcular Poder del Uno"):
        _resu = comparativo_estados(cli_id, mes_pu)
        if not _resu:
            st.error("No se encontró ese mes.")
        else:
            _efu, _efpu, _ = _resu
            _baseu = base_poder_uno(_efu, _efpu, mes_pu[:7], temp_pu)
            _mvu = dict(precio=pu_precio, volumen=pu_vol, costo=pu_costo, gastos=pu_gastos, cxc=pu_cxc, inv=pu_inv, cxp=pu_cxp)
            _filas, _du, _dc, _trampa, _ub = poder_uno_tabla(_baseu, _mvu)
            st.session_state["pu_scenario"] = dict(temporalidad=temp_pu, filas=_filas, du=_du, dc=_dc,
                                                   trampa=_trampa, periodo=mes_pu[:7])
            st.dataframe([{"Palanca": n, "Movimiento": mv,
                           "Efecto": ("+" if v >= 0 else "-") + money(abs(v)), "Sobre": k} for n, mv, v, k in _filas],
                         use_container_width=True, hide_index=True)
            cc1, cc2 = st.columns(2)
            cc1.metric("Cambio en utilidad", ("+" if _du >= 0 else "-") + money(abs(_du)))
            cc2.metric("Cambio en caja", ("+" if _dc >= 0 else "-") + money(abs(_dc)))
            if _trampa:
                st.error("⚠️ Trampa de volumen: con margen bruto negativo (" + money(_ub) + "), vender más volumen REDUCE la utilidad. "
                         "El sistema no lo maquilla: primero el margen, después el volumen.")
            st.caption("Escenario guardado. Al generar el reporte del cliente de " + mes_pu[:7] + " se incluye esta página.")


# ---------------------------------------------------------------------------
# REPORTE DEL CLIENTE (vista dueño)
# ---------------------------------------------------------------------------
st.divider()
st.subheader("Reporte del cliente")
_periodos = periodos_cargados(cli_id)
if not _periodos:
    st.caption("Carga al menos un mes para generar el reporte del cliente.")
else:
    mes_rep = st.selectbox("Mes a reportar", _periodos, key="rep_mes")
    if st.button("Generar reporte del cliente"):
        st.session_state["rep_generado"] = mes_rep
    if st.session_state.get("rep_generado") == mes_rep:
        ind = datos_reporte_cliente(cli_id, mes_rep)
        if ind is None:
            st.error("No se encontró ese mes.")
        else:
            st.markdown(f"### {nombre_sel}  ·  {mes_rep[:7]}")
            # caja al centro
            if ind["caja_var"] is not None:
                st.metric("Tu caja al cierre del mes", money(ind["caja"]), delta=f"{ind['caja_var']:,.0f} vs mes anterior")
                direccion = "bajó" if ind["caja_var"] < 0 else "subió"
                st.write(f"Tu caja {direccion} **{money(abs(ind['caja_var']))}** respecto al mes anterior.")
            else:
                st.metric("Tu caja al cierre del mes", money(ind["caja"]))
                st.caption("Sin mes anterior para comparar.")
            # tres indicadores en lenguaje de dueño
            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric("Rentabilidad del año (acumulada)", f"{ind['pretax_pct']}%" if ind["pretax_pct"] is not None else "—")
                st.caption(sem_pretax(ind["pretax_pct"]))
            with c2:
                st.metric("Margen de tu operación", f"{ind['mb_pct']}%" if ind["mb_pct"] is not None else "—")
            with c3:
                st.metric("Brecha utilidad vs. caja", money(ind["cash_lag"]))
                _uai = ind.get("uai"); _cl = ind.get("cash_lag")
                if _cl is None or _uai is None:
                    st.caption("Sin dato")
                elif _uai <= 0:
                    st.caption("⚪ Hay pérdida; la prioridad es la rentabilidad")
                elif _cl > 0:
                    st.caption("🔴 La utilidad no llegó a caja")
                else:
                    st.caption("🟢 La caja siguió a la utilidad")
            # lectura del mes (la escribe Roberto)
            st.markdown("**Lo que esto significa** (escríbelo para el dueño)")
            st.text_area("Lectura del mes", key="rep_lectura",
                         placeholder="Ej.: El negocio es rentable, pero la utilidad se quedó en cobranza. Prioridad: cobrar, no vender.",
                         label_visibility="collapsed")
            st.caption("Los números los pone el sistema. La lectura la escribe el CFO.")
            st.text_input("El número más importante del mes (opcional)", key="rep_numero",
                          placeholder="Ej.: Tu caja subió $1.7M, pero $9M siguen en cobranza.")
            st.text_area("Las 3 acciones del mes (una por línea)", key="rep_acciones",
                         placeholder="Cobrar la cartera vencida de obra\nFrenar compra de material sin contrato firmado\nReclasificar el crédito de corto a largo plazo")
            st.text_area("Valor generado este mes (en pesos)", key="rep_valor",
                         placeholder="Ej.: Recuperamos $420,000 de IVA a favor. Acumulado del año: $1.3M.")
            try:
                import fpdf  # noqa: F401
                _fpdf_ok = True
            except Exception:
                _fpdf_ok = False
            if not _fpdf_ok:
                st.error("Falta la librería **fpdf2**. En GitHub: agrega una línea `fpdf2` a requirements.txt, "
                         "guarda, y en Streamlit Cloud entra a *Manage app* y haz *Reboot*. Sin esto no se genera el PDF.")
            else:
                try:
                    _res_ef = comparativo_estados(cli_id, mes_rep)
                    _ef_a, _ef_p, _per_p = _res_ef if _res_ef else (None, None, None)
                    _rat = ratios_mensuales(cli_id, mes_rep)
                    _metas_c = get_metas(cli_id, int(mes_rep[:4]))
                    _pu = st.session_state.get("pu_scenario")
                    if _pu and _pu.get("periodo") != mes_rep[:7]:
                        _pu = None
                    _pdf = pdf_reporte_cliente(nombre_sel, mes_rep[:7], ind, st.session_state.get("rep_lectura", ""),
                                               ef_a=_ef_a, ef_p=_ef_p, per_p=_per_p, rat=_rat,
                                               numero_mes=st.session_state.get("rep_numero", ""),
                                               acciones=st.session_state.get("rep_acciones", ""),
                                               valor_generado=st.session_state.get("rep_valor", ""),
                                               metas=_metas_c, poder_uno=_pu, branding=get_branding(cli_id))
                    st.download_button("📄 Descargar PDF para el cliente", data=_pdf,
                                       file_name="Reporte_CFO_" + nombre_sel.replace(" ", "_") + "_" + mes_rep[:7] + ".pdf",
                                       mime="application/pdf", key="rep_pdf_dl")
                    st.caption("El PDF toma la lectura escrita arriba. Si la editas, haz clic fuera del cuadro antes de descargar.")
                except Exception as _e:
                    st.error("No se pudo generar el PDF: " + repr(_e))
